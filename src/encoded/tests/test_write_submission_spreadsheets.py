import tempfile
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Dict, List, Optional
from unittest import mock

import openpyxl
import pytest
from dcicutils import schema_utils
from dcicutils.misc_utils import to_camel_case, to_snake_case

from ..commands.write_submission_spreadsheets import (
    DROPDOWN_HELPER_SHEET_NAME,
    FONT,
    FONT_SIZE,
    ITEM_SPREADSHEET_SUFFIX,
    LOCATIONS_HELPER_SHEET_NAME,
    OVERVIEW_SHEET_NAME,
    WORKBOOK_FILENAME,
    Property,
    Spreadsheet,
    get_array_subtype,
    get_comment_text,
    get_enum,
    get_existing_static_content_identifier,
    get_font,
    get_ordered_properties,
    get_property,
    get_nested_properties,
    get_spreadsheet,
    is_link,
    write_all_spreadsheets,
    write_item_spreadsheets,
    write_publication_static_section_workbook,
)
from ..item_utils.utils import RequestHandler


@pytest.fixture
def submission_schema() -> Dict[str, Any]:
    """Mock submission schema for tests."""
    return {
        "title": "Foo",
        "properties": {
            "bar": {"title": "Bar", "type": "string"},
            "baz": {"title": "Baz", "type": "number"},
        },
    }


@pytest.fixture
def submission_schemas(
    submission_schema: Dict[str, Any],
) -> Dict[str, Dict[str, Any]]:
    """Mock response from /submission-schemas/ for tests."""
    return {
        "Foo": submission_schema,
        "Qux": {
            "title": "Qux",
            "properties": {
                "quux": {"title": "Quux", "type": "string"},
                "corge": {"title": "Corge", "type": "number"},
            },
        },
    }


def get_mock_request_handler() -> mock.Mock:
    return mock.create_autospec(RequestHandler)


@contextmanager
def patch_get_all_submission_schemas(
    submission_schemas: Dict[str, Dict[str, Any]],
) -> mock.Mock:
    with mock.patch(
        "encoded.commands.write_submission_spreadsheets.get_all_submission_schemas",
        return_value=submission_schemas,
    ) as mock_get_all_submission_schemas:
        yield mock_get_all_submission_schemas


@contextmanager
def patch_get_submission_schemas(
    submission_schemas: Dict[str, Dict[str, Any]],
) -> mock.Mock:
    with mock.patch(
        "encoded.commands.write_submission_spreadsheets.get_submission_schemas",
        return_value=submission_schemas,
    ) as mock_get_submission_schemas:
        yield mock_get_submission_schemas


def get_item_index_order(submission_schemas: Dict[str, Dict[str, Any]]) -> List[str]:
    """Create mock ITEM_INDEX_ORDER based on submission schemas.

    In this case, determined by reverse alphabetical order of item types.
    """
    sorted_order = sorted(
        [to_snake_case(item_type) for item_type in submission_schemas.keys()]
    )
    return sorted_order[::-1]


@contextmanager
def patch_item_index_order(submission_schemas: Dict[str, Dict[str, Any]]) -> mock.Mock:
    new_item_index_order = get_item_index_order(submission_schemas)
    with mock.patch(
        "encoded.commands.write_submission_spreadsheets.ITEM_INDEX_ORDER",
        new_item_index_order,
    ) as mock_item_index_order:
        yield mock_item_index_order


@pytest.mark.parametrize(
    "workbook,separate_comments",
    [
        (False, False),
        (False, True),
        (True, False),
        (True, True),
        (True, False),
    ],
)
def test_write_all_spreadsheets(
    workbook: bool,
    separate_comments: bool,
    submission_schemas: Dict[str, Dict[str, Any]],
) -> None:
    """Test writing all spreadsheets.

    Serves as an integrated test for high-level functionality. Details
    tested in other unit tests.
    """
    request_handler = get_mock_request_handler()
    with tempfile.TemporaryDirectory() as tempdir:
        with patch_get_all_submission_schemas(submission_schemas):
            with patch_item_index_order(submission_schemas):
                write_all_spreadsheets(
                    tempdir,
                    request_handler,
                    workbook=workbook,
                    separate_comments=separate_comments,
                )
        if workbook:
            assert_workbook_written(tempdir, submission_schemas, separate_comments)
        else:
            assert_spreadsheets_written(tempdir, submission_schemas, separate_comments)


@pytest.mark.parametrize(
    "items,workbook,separate_comments,expected_items",
    [
        ([], False, False, []),
        (["Fu"], True, True, []),
        (["Foo"], False, False, ["Foo"]),
        (["Foo"], False, True, ["Foo"]),
        (["Foo"], True, False, ["Foo"]),
        (["Foo"], True, True, ["Foo"]),
        (["Foo", "Qux"], False, False, ["Foo", "Qux"]),
        (["Foo", "Fu"], False, False, ["Foo"]),
    ],
)
def test_write_item_spreadsheets(
    items: List[str],
    workbook: bool,
    separate_comments: bool,
    expected_items: List[str],
    submission_schemas: Dict[str, Dict[str, Any]],
) -> None:
    """Test writing spreadsheets for given item types."""
    request_handler = get_mock_request_handler()
    with tempfile.TemporaryDirectory() as tempdir:
        expected_submission_schemas = get_expected_schemas(
            submission_schemas, expected_items
        )
        with patch_get_submission_schemas(expected_submission_schemas):
            with patch_item_index_order(submission_schemas):
                write_item_spreadsheets(
                    tempdir,
                    items,
                    request_handler,
                    workbook=workbook,
                    separate_comments=separate_comments,
                )
        if not expected_items:
            assert list(Path(tempdir).iterdir()) == []
        else:
            if workbook:
                assert_workbook_written(
                    tempdir, submission_schemas, separate_comments, items=expected_items
                )
            else:
                assert_spreadsheets_written(
                    tempdir, submission_schemas, separate_comments, items=expected_items
                )


def assert_workbook_written(
    tempdir: str,
    submission_schemas: Dict[str, Dict[str, Any]],
    separate_comments: bool,
    items: Optional[List[str]] = None
) -> None:
    """Assert that the workbook was written correctly at a high level."""
    workbook_path = Path(tempdir).joinpath(WORKBOOK_FILENAME)
    assert workbook_path.exists()
    workbook = openpyxl.load_workbook(workbook_path)
    expected_schemas = get_expected_schemas(submission_schemas, items)
    assert_workbook_sheets_ordered(workbook, expected_schemas)
    for item_type, schema in expected_schemas.items():
        sheet = workbook.get_sheet_by_name(item_type)
        assert_sheet_written(sheet, schema, separate_comments)


def get_expected_schemas(
    submission_schemas: Dict[str, Dict[str, Any]],
    items: Optional[List[str]] = None
) -> Dict[str, Dict[str, Any]]:
    """Get the expected schemas for the given items."""
    if items is None:
        return submission_schemas
    return {
        item_type: submission_schemas.get(item_type)
        for item_type in items
        if item_type in submission_schemas
    }


def assert_workbook_sheets_ordered(
    workbook: openpyxl.workbook.workbook.Workbook,
    expected_schemas: Dict[str, Dict[str, Any]],
) -> None:
    """Assert that the workbook sheets are ordered correctly."""
    expected_sheet_names = [
        to_camel_case(item_type) for item_type in get_item_index_order(expected_schemas)
    ]
    assert workbook.sheetnames == expected_sheet_names


def assert_sheet_written(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    schema: Dict[str, Any],
    separate_comments: bool,
) -> None:
    """Assert that a sheet was written correctly at a high level.

    Not checking cell values here, just the structure.
    """
    expected_properties = schema_utils.get_properties(schema)
    expected_property_names = set(expected_properties.keys())
    first_row_cells = [cell for cell in sheet[1]]
    second_row_cell_values = [cell.value for cell in sheet[2]]
    actual_property_names = set(cell.value for cell in first_row_cells)
    assert actual_property_names == expected_property_names
    if separate_comments:
        assert len(second_row_cell_values) == len(actual_property_names)
        for cell_value in second_row_cell_values:
            assert cell_value
        for cell in first_row_cells:
            assert not cell.comment
    else:
        assert set(second_row_cell_values) == {None}
        for cell in first_row_cells:
            assert cell.comment


def assert_spreadsheets_written(
    tempdir: str,
    submission_schemas: Dict[str, Dict[str, Any]],
    separate_comments: bool,
    items: Optional[List[str]] = None,
) -> None:
    """Assert that spreadsheets were written correctly at a high level."""
    expected_schemas = get_expected_schemas(submission_schemas, items)
    for item_type, schema in expected_schemas.items():
        spreadsheet_path = Path(tempdir).joinpath(
            f"{to_snake_case(item_type)}{ITEM_SPREADSHEET_SUFFIX}"
        )
        assert spreadsheet_path.exists()
        workbook = openpyxl.load_workbook(spreadsheet_path)
        assert workbook.sheetnames == [item_type]
        sheet = workbook.get_sheet_by_name(item_type)
        assert_sheet_written(sheet, schema, separate_comments)


def test_get_spreadsheet(submission_schema: Dict[str, Any]) -> None:
    """Test creation of Spreadsheet class from schema."""
    item = "foo"
    spreadsheet = get_spreadsheet(item, submission_schema)
    assert isinstance(spreadsheet, Spreadsheet)
    assert spreadsheet.item == item
    assert len(spreadsheet.properties) == len(
        schema_utils.get_properties(submission_schema)
    )
    for property_ in spreadsheet.properties:
        assert isinstance(property_, Property)


@pytest.mark.parametrize(
    "property_name,property_schema,expected",
    [
        (  # Simple case with defaults
            "bar",
            {},
            Property(
                name="bar",
                item="Foo",
                description="",
                value_type="",
                required=False,
                link=False,
                enum=[],
                array_subtype="",
                pattern="",
                comment="",
                examples=[],
                format_="",
                requires=[],
                exclusive_requirements=[],
            ),
        ),
        (  # More complicated case with most attributes
            "baz",
            {
                "description": "Baz",
                "type": "number",
                "is_required": True,
                "linkTo": "Bar",
                "enum": [1, 2, 3],
                "pattern": "pattern",
                "submissionComment": "This is a comment.",
                "submissionExamples": [1, 2],
                "format": "format",
                "also_requires": ["foo"],
                "required_if_not_one_of": ["bar"],
            },
            Property(
                name="baz",
                item="Foo",
                description="Baz",
                value_type="number",
                required=True,
                link=True,
                enum=[1, 2, 3],
                array_subtype="",
                pattern="pattern",
                comment="This is a comment.",
                examples=[1, 2],
                format_="format",
                requires=["foo"],
                exclusive_requirements=["bar"],
                search="https://data.smaht.org/search/?type=Bar"
            ),
        ),
        (  # More complicated case with suggested_enum
            "baz",
            {
                "description": "Baz",
                "type": "number",
                "is_required": True,
                "linkTo": "Bar",
                "enum": [1, 2, 3],
                "pattern": "pattern",
                "submissionComment": "This is a comment.",
                "suggested_enum": [1, 2],
                "format": "format",
                "also_requires": ["foo"],
                "required_if_not_one_of": ["bar"],
            },
            Property(
                name="baz",
                item="Foo",
                description="Baz",
                value_type="number",
                required=True,
                link=True,
                enum=[1, 2, 3],
                array_subtype="",
                pattern="pattern",
                comment="This is a comment.",
                examples=[1, 2],
                format_="format",
                requires=["foo"],
                exclusive_requirements=["bar"],
                search="https://data.smaht.org/search/?type=Bar"
            ),
        ),
    ],
)
def test_get_property(
    property_name: str, property_schema: Dict[str, Any], expected: Property
) -> None:
    """Test creation of Property class from schema.

    For more complicated attributes, see respective unit tests.
    """
    property_ = get_property('Foo',property_name, property_schema)
    assert property_ == expected


@pytest.mark.parametrize(
    "property_name,property_schema,expected",
    [
        (
            "baz", # test array of objects
            {
                "description": "Baz",
                "type": "array",
                "is_required": True,
                "items": {
                    "properties": {
                        "foo": {
                            "description": "Foo",
                            "type": "number"
                        }
                    }
                }
            },
            [
                Property(
                    name="baz#0.foo",
                    item="Foo",
                    description="Foo",
                    value_type="number",
                    required=False,
                    link=False,
                    enum=[],
                    array_subtype="",
                    pattern="",
                    comment="",
                    examples=[],
                    format_="",
                    requires=[],
                    exclusive_requirements=[],
                    nested=True
                ),
                Property(
                    name="baz#1.foo",
                    item="Foo",
                    description="Foo",
                    value_type="number",
                    required=False,
                    link=False,
                    enum=[],
                    array_subtype="",
                    pattern="",
                    comment="",
                    examples=[],
                    format_="",
                    requires=[],
                    exclusive_requirements=[],
                    nested=True
                )
            ]
        ),
    ]
)
def test_get_nested_properties(
    property_name: str, property_schema: Dict[str, Any], expected: Property
) -> None:
    """Test get_nested_properties from schema.
    """
    property_ = get_nested_properties("Foo",property_name, property_schema)
    assert property_ == expected


def test_get_nested_properties_plain_object() -> None:
    """Test that a plain (non-array) nested object is flattened to `parent.child` columns."""
    property_schema = {
        "type": "object",
        "properties": {
            "filetype": {"type": "string", "enum": ["md", "html"]},
            "collapsible": {"type": "boolean"},
        },
    }
    properties = get_nested_properties("StaticSection", "options", property_schema)
    assert [property_.name for property_ in properties] == [
        "options.filetype", "options.collapsible",
    ]
    assert all(property_.item == "StaticSection" for property_ in properties)
    assert all(not property_.nested for property_ in properties)
    assert all(not property_.required for property_ in properties)
    filetype_property = properties[0]
    assert filetype_property.enum == ["md", "html"]


def test_get_flattened_object_properties_required_propagation() -> None:
    """Test that a plain nested object's own `required` list marks the matching child columns required."""
    property_schema = {
        "type": "object",
        "required": ["filetype"],
        "properties": {
            "filetype": {"type": "string"},
            "collapsible": {"type": "boolean"},
        },
    }
    properties = get_nested_properties("StaticSection", "options", property_schema)
    by_name = {property_.name: property_ for property_ in properties}
    assert by_name["options.filetype"].required is True
    assert by_name["options.collapsible"].required is False


@pytest.mark.parametrize(
    "array_slot_counts,default_array_slot_count,expected_count",
    [
        (None, 2, 2),  # unchanged default, matches legacy hard-coded behavior
        (None, 5, 5),  # global override
        ({"baz": 4}, 2, 4),  # per-property override takes precedence over the default
        ({"other": 9}, 2, 2),  # override for a different property name doesn't apply
    ],
)
def test_get_nested_properties_configurable_slot_count(
    array_slot_counts: Optional[Dict[str, int]],
    default_array_slot_count: int,
    expected_count: int,
) -> None:
    """Test that the array-of-objects slot count is configurable per-property or globally."""
    property_schema = {
        "type": "array",
        "items": {"properties": {"foo": {"type": "number"}}},
    }
    properties = get_nested_properties(
        "Foo",
        "baz",
        property_schema,
        array_slot_counts=array_slot_counts,
        default_array_slot_count=default_array_slot_count,
    )
    assert [property_.name for property_ in properties] == [
        f"baz#{index}.foo" for index in range(expected_count)
    ]


def test_get_nested_properties_array_required_propagation() -> None:
    """Test that an array-of-objects `items.required` list marks the matching slot columns required."""
    property_schema = {
        "type": "array",
        "items": {
            "required": ["content", "location"],
            "properties": {
                "content": {"type": "string", "linkTo": "UserContent"},
                "location": {"type": "string"},
                "description": {"type": "string"},
            },
        },
    }
    properties = get_nested_properties(
        "Publication", "static_content", property_schema, array_slot_counts={"static_content": 2}
    )
    by_name = {property_.name: property_ for property_ in properties}
    assert by_name["static_content#0.content"].required is True
    assert by_name["static_content#0.location"].required is True
    assert by_name["static_content#0.description"].required is False
    assert by_name["static_content#1.content"].required is True
    assert by_name["static_content#1.location"].required is True
    assert by_name["static_content#1.description"].required is False


@pytest.mark.parametrize(
    "property_schema,expected",
    [
        ({}, False),
        ({"linkTo": "Foo"}, True),
        ({"type": "array", "items": {"linkTo": "Foo"}}, True),
        ({"type": "array", "items": {"type": "string"}}, False),
    ],
)
def test_is_link(property_schema: Dict[str, Any], expected: bool) -> None:
    """Test determination of whether a property is a link."""
    assert is_link(property_schema) == expected


@pytest.mark.parametrize(
    "property_schema,expected",
    [
        ({}, []),
        ({"enum": [1, 2, 3]}, [1, 2, 3]),
        ({"type": "array", "items": {"enum": [1, 2, 3]}}, [1, 2, 3]),
        ({"type": "array", "items": {"type": "string"}}, []),
    ],
)
def test_get_enum(property_schema: Dict[str, Any], expected: List[str]) -> None:
    """Test retrieval of enum from property schema."""
    assert get_enum(property_schema) == expected


@pytest.mark.parametrize(
    "property_schema,expected",
    [
        ({}, ""),
        ({"type": "array", "items": {"type": "string"}}, "string"),
        ({"type": "array", "items": {"type": "number"}}, "number"),
        (  # Even though not handling arrays of objects currently -DRR 2024-06-10
            {
                "type": "array",
                "items": {"type": "object", "properties": {"foo": {"type": "string"}}},
            },
            "object",
        ),
    ],
)
def test_get_array_subtype(property_schema: Dict[str, Any], expected: str) -> None:
    """Test retrieval of array subtype from property schema."""
    assert get_array_subtype(property_schema) == expected


@pytest.mark.parametrize(
    "properties,expected",
    [
        ([], []),
        (
            [
                Property("bar"),
                Property("baa"),
                Property("tiz", required=True),
                Property("foo", required=True),
                Property("submitted_id", required=True),
                Property("qux", link=True),
                Property("baz", link=True),
                Property("quux", link=True, required=True),
                Property("corge", link=True, required=True),
            ],
            [
                Property("submitted_id", required=True),
                Property("foo", required=True),
                Property("tiz", required=True),
                Property("bar"),
                Property("baa"),
                Property("quux", link=True, required=True),
                Property("corge", link=True, required=True),
                Property("baz", link=True),
                Property("qux", link=True),
            ],
        ),
    ],
)
def test_get_ordered_properties(
    properties: List[Property], expected: List[Property]
) -> None:
    """Test that properties are ordered correctly."""
    assert get_ordered_properties(properties) == expected


@pytest.mark.parametrize(
    "property_,expected",
    [
        (Property("foo"), "Required:  No"),  # Base case
        (  # Most attributes
            Property(
                "foo",
                description="Foo",
                required=True,
                link=True,
                value_type="string",
                enum=["foo", "bar"],
                examples=["fu", "bar"],
                requires=["bar", "bu"],
                pattern="pattern",
                comment="This is a comment.",
            ),
            (
                "Description:  Foo\n"
                "Type:  string\n"
                "Options:  foo | bar\n"
                "Examples:  fu | bar\n"
                "Link:  Yes\n"
                "Required:  Yes\n"
                "Requires:  bar | bu\n"
                "Pattern:  pattern\n"
                "Note:  This is a comment."
            ),
        ),
        (  # Array type
            Property("foo", value_type="array", array_subtype="string"),
            "Type:  string  (Multiple values allowed. Use '|' as a delimiter.)\nRequired:  No",
        ),
        (  # Possibly required
            Property("foo", exclusive_requirements=["bar", "bu"]),
            "Required:  Possibly\n  Not required if present:  bar | bu",
        ),
        (  # Date format without pattern
            Property("foo", format_="date"),
            "Required:  No\nFormat:  YYYY-MM-DD",
        ),
        (  # Most search and nested
            Property(
                "foo",
                description="Foo",
                required=False,
                link=True,
                value_type="string",
                nested=True,
                search="test.url.com",
            ),
            (
                "Description:  Foo\n"
                "Type:  string\n"
                "Link:  Yes\n"
                "Required:  No\n"
                "Use URL to search for the submitted_id or identifer of relevant items:  test.url.com\n"
                "Nested:  Yes"
            ),
        ),
    ],
)
def test_get_comment_text(property_: Property, expected: str) -> None:
    """Test that comments are generated correctly."""
    assert get_comment_text(property_) == expected


@pytest.mark.parametrize(
    "property_,expected_bold,expected_italic",
    [
        (
            Property("foo"),
            False,
            False,
        ),
        (
            Property("foo", required=True),
            True,
            False,
        ),
        (
            Property("foo", link=True),
            False,
            True,
        ),
    ],
)
def test_get_font(
    property_: Property, expected_bold: bool, expected_italic: bool
) -> None:
    """Test that font is generated correctly."""
    result = get_font(property_)
    assert isinstance(result, openpyxl.styles.Font)
    assert result.name == FONT
    assert result.size == FONT_SIZE
    if expected_bold:
        assert result.bold
    else:
        assert not result.bold
    if expected_italic:
        assert result.italic
    else:
        assert not result.italic


@pytest.fixture
def raw_static_section_schema() -> Dict[str, Any]:
    """Mock raw (non-submission-filtered) /profiles/static_section.json schema.

    Shaped like the real schema (see src/encoded/schemas/static_section.json and
    encoded_core's static_section.json): retains `required`, `submission_centers`,
    `consortia`, and `uuid`, none of which the /submission-schemas/ endpoint
    would carry.
    """
    return {
        "required": ["identifier"],
        "properties": {
            "identifier": {
                "title": "Identifier",
                "type": "string",
                "pattern": "^([A-Za-z0-9-_]+[.])*[A-Za-z0-9-_]+$",
            },
            "title": {"title": "Title", "type": "string"},
            "body": {"title": "Raw Body", "type": "string", "comment": "There should be no 'file' if this is set."},
            "file": {"title": "Source File Location", "type": "string", "comment": "There should be no 'body' if this is set."},
            "section_type": {
                "title": "Section Type",
                "type": "string",
                "default": "Page Section",
                "enum": ["Page Section", "Announcement", "Search Info Header", "Item Page Header", "Home Page Slide"],
            },
            "options": {
                "title": "Options",
                "type": "object",
                "properties": {
                    "filetype": {"type": "string", "enum": ["md", "html", "txt", "csv", "jsx", "rst"]},
                    "collapsible": {"type": "boolean", "default": False},
                    "default_open": {"type": "boolean", "default": True},
                    "title_icon": {"type": "string"},
                    "link": {"type": "string"},
                    "image": {"type": "string"},
                },
            },
            "description": {"title": "Description", "type": "string"},
            "submission_centers": {
                "title": "Generated By", "type": "array", "items": {"type": "string", "linkTo": "SubmissionCenter"},
            },
            "consortia": {
                "title": "Consortia", "type": "array", "items": {"type": "string", "linkTo": "Consortium"},
            },
            "uuid": {"title": "UUID", "type": "string"},
            "aliases": {"title": "Aliases", "type": "array", "items": {"type": "string"}},
            "status": {"title": "Status", "type": "string"},
        },
    }


@pytest.fixture
def raw_publication_schema() -> Dict[str, Any]:
    """Mock raw (non-submission-filtered) /profiles/publication.json schema.

    Retains `accession`, `uuid`, and the nested `items.required` for
    `static_content`, none of which the /submission-schemas/ endpoint would
    carry (see src/encoded/schemas/publication.json and mixins.json#/static_embeds).
    """
    return {
        "required": ["doi"],
        "properties": {
            "accession": {"title": "Accession", "type": "string"},
            "uuid": {"title": "UUID", "type": "string"},
            "doi": {"title": "DOI", "type": "string"},
            "title": {"title": "Title", "type": "string"},
            "static_content": {
                "title": "Static Content",
                "type": "array",
                "items": {
                    "title": "Static Content Definition",
                    "type": "object",
                    "required": ["location", "content"],
                    "properties": {
                        "content": {"type": "string", "linkTo": "UserContent"},
                        "location": {"type": "string", "default": "header"},
                        "description": {"type": "string"},
                    },
                },
            },
            "static_headers": {
                "title": "Static Headers", "type": "array", "items": {"type": "string", "linkTo": "UserContent"},
            },
        },
    }


@contextmanager
def patch_publication_static_section_profiles(
    raw_static_section_schema: Dict[str, Any],
    raw_publication_schema: Dict[str, Any],
    static_sections_by_id: Optional[Dict[str, Dict[str, Any]]] = None,
) -> mock.Mock:
    """Patch `RequestHandler.get_item` to serve the mock raw profile schemas.

    `static_sections_by_id`, if given, additionally resolves specific StaticSection
    `@id`s via `ff_utils.get_metadata` (as populated-mode static_content pre-fill,
    via `get_existing_static_content_identifier`, would look up).
    """
    static_sections_by_id = static_sections_by_id or {}

    def fake_get_item(identifier: str, collection: Optional[str] = None) -> Dict[str, Any]:
        if "static_section" in identifier:
            return raw_static_section_schema
        if "publication" in identifier:
            return raw_publication_schema
        return {}

    def fake_get_metadata(identifier: str, key: Optional[Dict[str, str]] = None, add_on: str = "") -> Dict[str, Any]:
        return static_sections_by_id.get(identifier, {})

    request_handler = get_mock_request_handler()
    request_handler.get_item.side_effect = fake_get_item
    request_handler.auth_key = {"server": "https://example.com", "key": "test-key", "secret": "test-secret"}
    with mock.patch(
        "encoded.commands.write_submission_spreadsheets.ff_utils.get_metadata", side_effect=fake_get_metadata
    ):
        yield request_handler


def get_cell_value_by_header(sheet, header: str, row: int) -> Any:
    """Get a data-row cell's value given its header name, for readability in assertions."""
    headers = [cell.value for cell in sheet[1]]
    column = headers.index(header) + 1
    return sheet.cell(row=row, column=column).value


def test_write_publication_static_section_workbook_template_mode(
    raw_static_section_schema: Dict[str, Any], raw_publication_schema: Dict[str, Any]
) -> None:
    """Test the Publication/StaticSection workbook's structure, ordering, and formatting in template mode."""
    with patch_publication_static_section_profiles(
        raw_static_section_schema, raw_publication_schema
    ) as request_handler:
        with tempfile.TemporaryDirectory() as tempdir:
            output = Path(tempdir, "test.xlsx")
            write_publication_static_section_workbook(
                output, request_handler, static_content_slots=2, locations=["key-findings", "header"]
            )
            workbook = openpyxl.load_workbook(output)

            # Sheet order and visibility.
            assert workbook.sheetnames == [
                OVERVIEW_SHEET_NAME,
                "StaticSection",
                "Publication",
                DROPDOWN_HELPER_SHEET_NAME,
                LOCATIONS_HELPER_SHEET_NAME,
            ]
            assert workbook[OVERVIEW_SHEET_NAME].sheet_state == "visible"
            assert workbook["StaticSection"].sheet_state == "visible"
            assert workbook["Publication"].sheet_state == "visible"
            assert workbook[DROPDOWN_HELPER_SHEET_NAME].sheet_state == "hidden"
            assert workbook[LOCATIONS_HELPER_SHEET_NAME].sheet_state == "hidden"

            static_section_sheet = workbook["StaticSection"]
            static_section_headers = [cell.value for cell in static_section_sheet[1]]
            assert set(static_section_headers) == {
                "identifier", "title", "body", "file", "section_type",
                "options.filetype", "options.collapsible", "options.default_open",
                "options.title_icon", "options.link", "options.image",
                "description", "submission_centers", "consortia",
            }
            # No data rows in template mode.
            assert all(cell.value is None for cell in static_section_sheet[2])

            publication_sheet = workbook["Publication"]
            publication_headers = [cell.value for cell in publication_sheet[1]]
            assert set(publication_headers) == {
                "accession", "uuid",
                "static_content#0.content", "static_content#0.location", "static_content#0.description",
                "static_content#1.content", "static_content#1.location", "static_content#1.description",
                "static_headers",
            }
            assert all(cell.value is None for cell in publication_sheet[2])

            # Bold (required) / italic (link) formatting.
            def get_header_cell(sheet, name):
                headers = [cell for cell in sheet[1]]
                return next(cell for cell in headers if cell.value == name)

            assert get_header_cell(static_section_sheet, "identifier").font.bold
            assert not get_header_cell(static_section_sheet, "identifier").font.italic
            assert get_header_cell(static_section_sheet, "submission_centers").font.italic
            assert not get_header_cell(static_section_sheet, "title").font.bold

            assert get_header_cell(publication_sheet, "accession").font.bold
            assert get_header_cell(publication_sheet, "static_content#0.content").font.bold
            assert get_header_cell(publication_sheet, "static_content#0.content").font.italic
            assert get_header_cell(publication_sheet, "static_content#0.location").font.bold
            assert not get_header_cell(publication_sheet, "static_content#0.location").font.italic
            assert not get_header_cell(publication_sheet, "static_content#0.description").font.bold
            assert not get_header_cell(publication_sheet, "uuid").font.bold

            # Dropdowns: StaticSection gets 4 (section_type, filetype, collapsible, default_open);
            # Publication gets 1 per slot (location only, not content).
            assert len(static_section_sheet.data_validations.dataValidation) == 4
            assert len(publication_sheet.data_validations.dataValidation) == 2
            location_validation_formulas = {
                dv.formula1 for dv in publication_sheet.data_validations.dataValidation
            }
            assert location_validation_formulas == {f"='{LOCATIONS_HELPER_SHEET_NAME}'!$A$2:$A$3"}

            # Helper sheets carry the enum/location source data.
            dropdown_sheet = workbook[DROPDOWN_HELPER_SHEET_NAME]
            assert dropdown_sheet["A1"].value == "Section Type"
            assert dropdown_sheet["A2"].value == "Page Section"
            assert dropdown_sheet["B1"].value == "File Type"
            assert dropdown_sheet["C1"].value == "Boolean"
            locations_sheet = workbook[LOCATIONS_HELPER_SHEET_NAME]
            assert locations_sheet["A2"].value == "key-findings"
            assert locations_sheet["A3"].value == "header"


def test_write_publication_static_section_workbook_populated_append(
    raw_static_section_schema: Dict[str, Any], raw_publication_schema: Dict[str, Any]
) -> None:
    """Test populated mode pre-fills accession/uuid and existing static_content when appending.

    The existing entry's linked `content` is resolved to the target StaticSection's
    `identifier` (not left as the raw `@id`), so the pre-filled sheet is directly
    re-submittable.
    """
    existing_publications = [
        {
            "accession": "SMAHT001",
            "uuid": "11111111-1111-1111-1111-111111111111",
            "static_content": [
                {"content": "/static-sections/some-uuid/", "location": "key-findings"},
            ],
        },
    ]
    static_sections_by_id = {
        "/static-sections/some-uuid/": {"identifier": "SMAHT001.key-findings", "uuid": "some-uuid"},
    }
    with patch_publication_static_section_profiles(
        raw_static_section_schema, raw_publication_schema, static_sections_by_id
    ) as request_handler:
        with mock.patch(
            "encoded.commands.write_submission_spreadsheets.ff_utils.search_metadata",
            return_value=existing_publications,
        ):
            with tempfile.TemporaryDirectory() as tempdir:
                output = Path(tempdir, "test.xlsx")
                write_publication_static_section_workbook(
                    output,
                    request_handler,
                    static_content_slots=2,
                    locations=["key-findings", "header"],
                    mode="populated",
                    append=True,
                )
                workbook = openpyxl.load_workbook(output)
                publication_sheet = workbook["Publication"]
                assert get_cell_value_by_header(publication_sheet, "accession", 2) == "SMAHT001"
                assert get_cell_value_by_header(publication_sheet, "uuid", 2) == "11111111-1111-1111-1111-111111111111"
                assert (
                    get_cell_value_by_header(publication_sheet, "static_content#0.content", 2)
                    == "SMAHT001.key-findings"
                )
                assert get_cell_value_by_header(publication_sheet, "static_content#0.location", 2) == "key-findings"
                assert get_cell_value_by_header(publication_sheet, "static_content#1.content", 2) is None


def test_write_publication_static_section_workbook_populated_replace(
    raw_static_section_schema: Dict[str, Any], raw_publication_schema: Dict[str, Any]
) -> None:
    """Test populated mode still pre-fills accession/uuid but NOT static_content when replacing."""
    existing_publications = [
        {
            "accession": "SMAHT001",
            "uuid": "11111111-1111-1111-1111-111111111111",
            "static_content": [
                {"content": "/static-sections/SMAHT001.key-findings/", "location": "key-findings"},
            ],
        },
    ]
    with patch_publication_static_section_profiles(
        raw_static_section_schema, raw_publication_schema
    ) as request_handler:
        with mock.patch(
            "encoded.commands.write_submission_spreadsheets.ff_utils.search_metadata",
            return_value=existing_publications,
        ):
            with tempfile.TemporaryDirectory() as tempdir:
                output = Path(tempdir, "test.xlsx")
                write_publication_static_section_workbook(
                    output,
                    request_handler,
                    static_content_slots=2,
                    locations=["key-findings", "header"],
                    mode="populated",
                    append=False,
                )
                workbook = openpyxl.load_workbook(output)
                publication_sheet = workbook["Publication"]
                assert get_cell_value_by_header(publication_sheet, "accession", 2) == "SMAHT001"
                assert get_cell_value_by_header(publication_sheet, "static_content#0.content", 2) is None
                assert get_cell_value_by_header(publication_sheet, "static_content#0.location", 2) is None


def test_get_existing_static_content_identifier_resolves_linked_item() -> None:
    """Test that a static_content `@id` is resolved to the target StaticSection's identifier."""
    request_handler = get_mock_request_handler()
    request_handler.auth_key = {"server": "https://example.com", "key": "test-key", "secret": "test-secret"}
    with mock.patch(
        "encoded.commands.write_submission_spreadsheets.ff_utils.get_metadata",
        return_value={"identifier": "SMAHT001.key-findings", "uuid": "some-uuid"},
    ) as mock_get_metadata:
        result = get_existing_static_content_identifier("/static-sections/some-uuid/", request_handler)
    assert result == "SMAHT001.key-findings"
    mock_get_metadata.assert_called_once_with(
        "/static-sections/some-uuid/", key=request_handler.auth_key, add_on="frame=object"
    )


def test_get_existing_static_content_identifier_uses_portal_url_server() -> None:
    """Test that a given `portal_url` overrides the auth_key server used for resolution.

    This keeps the resolution lookup targeting the same server the Publication was
    fetched from, matching `get_existing_publications`'s own server override, rather
    than the server implied by `request_handler`'s `--env`-bound auth_key.
    """
    request_handler = get_mock_request_handler()
    request_handler.auth_key = {"server": "https://env-server.example.com", "key": "test-key", "secret": "test-secret"}
    with mock.patch(
        "encoded.commands.write_submission_spreadsheets.ff_utils.get_metadata",
        return_value={"identifier": "SMAHT001.key-findings", "uuid": "some-uuid"},
    ) as mock_get_metadata:
        result = get_existing_static_content_identifier(
            "/static-sections/some-uuid/", request_handler, portal_url="https://portal-url.example.com"
        )
    assert result == "SMAHT001.key-findings"
    mock_get_metadata.assert_called_once_with(
        "/static-sections/some-uuid/",
        key={"server": "https://portal-url.example.com", "key": "test-key", "secret": "test-secret"},
        add_on="frame=object",
    )


def test_get_existing_static_content_identifier_falls_back_to_raw_value() -> None:
    """Test that an unresolvable static_content `@id` falls back to the raw value rather than an empty cell."""
    request_handler = get_mock_request_handler()
    request_handler.auth_key = {"server": "https://example.com", "key": "test-key", "secret": "test-secret"}
    with mock.patch(
        "encoded.commands.write_submission_spreadsheets.ff_utils.get_metadata",
        return_value={},
    ):
        result = get_existing_static_content_identifier("/static-sections/some-uuid/", request_handler)
    assert result == "/static-sections/some-uuid/"


def test_get_existing_static_content_identifier_handles_empty_value() -> None:
    """Test that an empty static_content value is returned as-is without attempting a lookup."""
    request_handler = get_mock_request_handler()
    with mock.patch(
        "encoded.commands.write_submission_spreadsheets.ff_utils.get_metadata"
    ) as mock_get_metadata:
        assert get_existing_static_content_identifier("", request_handler) == ""
    mock_get_metadata.assert_not_called()
