import abc
import copy
import csv
from functools import lru_cache
import json
from jsonschema import Draft7Validator as SchemaValidator
import openpyxl
import os
import re
import sys
from typing import Any, Callable, Generator, Iterator, List, Optional, Tuple, Type, Union
from webtest.app import TestApp
from dcicutils.ff_utils import get_metadata, get_schema
from dcicutils.misc_utils import (merge_objects, remove_empty_properties, right_trim, split_string,
                                  to_boolean, to_camel_case, to_enum, to_float, to_integer, VirtualApp)
from dcicutils.zip_utils import temporary_file, unpack_gz_file_to_temporary_file, unpack_files
from snovault.loadxl import create_testapp

# Classes/functions to parse a CSV or Excel Spreadsheet into structured data, using a specialized
# syntax to allow structured object properties to be referenced by column specifiers. This syntax
# uses an (intuitive) dot notation to reference nested objects, and a (less intuitive) notation
# utilizing the "#" character to reference array elements. May also further coerce data types by
# consulting an optionally specified JSON schema.
#
# Alternate and semantically equivalent implementation of dcicutils.{sheet,bundle}_utils.
# Spare time exercise, with benefit of sheet_utils implementation experience.

ACCEPTABLE_FILE_SUFFIXES = [".csv", ".json", ".xls", ".xlsx", ".gz", ".tar", ".tar.gz", ".tgz", ".zip"]
ARRAY_VALUE_DELIMITER_CHAR = "|"
ARRAY_VALUE_DELIMITER_ESCAPE_CHAR = "\\"
ARRAY_NAME_SUFFIX_CHAR = "#"
ARRAY_NAME_SUFFIX_REGEX = re.compile(rf"{ARRAY_NAME_SUFFIX_CHAR}\d+")
DOTTED_NAME_DELIMITER_CHAR = "."

# Forward type references for type hints.
Portal = Type["Portal"]
PortalAny = Union[VirtualApp, TestApp, Portal]
RowReader = Type["RowReader"]
Schema = Type["Schema"]
StructuredDataSet = Type["StructuredDataSet"]


class StructuredDataSet:

    def __init__(self, file: Optional[str] = None, portal: Optional[PortalAny] = None,
                 schemas: Optional[List[dict]] = None, data: Optional[List[dict]] = None,
                 order: Optional[List[str]] = None, prune: bool = True) -> None:
        self.data = {} if not data else data
        self._portal = Portal.create(portal, data=self.data, schemas=schemas)  # If None then no schemas nor refs.
        self._order = order
        self._prune = prune
        self._issues = None
        self._load_file(file) if file else None

    @staticmethod
    def load(file: str, portal: Optional[PortalAny] = None, schemas: Optional[List[dict]] = None,
             order: Optional[List[str]] = None, prune: bool = True) -> StructuredDataSet:
        return StructuredDataSet(file=file, portal=portal, schemas=schemas, order=order, prune=prune)

    def validate(self) -> Optional[List[str]]:
        issues = []
        for type_name in self.data:
            if (schema := Schema.load_by_name(type_name, portal=self._portal)):
                for data in self.data[type_name]:
                    if (validate_issues := schema.validate(data)) is not None:
                        issues.extend(validate_issues)
        return issues + (self._issues or [])

    def _load_file(self, file: str) -> None:
        # Returns a dictionary where each property is the name (i.e. the type) of the data,
        # and the value is array of dictionaries for the data itself. Handle these kinds of files:
        # 1.  Single CSV of JSON file, where the (base) name of the file is the data type name.
        # 2.  Single Excel file containing one or more sheets, where each sheet
        #     represents (i.e. is named for, and contains data for) a different type.
        # 3.  Zip file (.zip or .tar.gz or .tgz or .tar), containing data files to load, where the
        #     base name of each contained file is the data type name; or any of above gzipped (.gz).
        if file.endswith(".gz") or file.endswith(".tgz"):
            with unpack_gz_file_to_temporary_file(file) as file:
                return self._load_normal_file(file)
        return self._load_normal_file(file)

    def _load_normal_file(self, file: str) -> None:
        if file.endswith(".csv"):
            self._load_csv_file(file)
        elif file.endswith(".xls") or file.endswith(".xlsx"):
            self._load_excel_file(file)
        elif file.endswith(".json"):
            self._load_json_file(file)
        elif file.endswith(".tar") or file.endswith(".zip"):
            self._load_packed_file(file)

    def _load_csv_file(self, file: str) -> None:
        reader = CsvReader(file)
        self._load_reader(reader, type_name=_get_type_name(file))
        self._note_issues(reader.issues, os.path.basename(file))

    def _load_excel_file(self, file: str) -> None:
        excel = Excel(file)  # Order the sheet names by any specified ordering (e.g. ala snovault.loadxl).
        order = {_get_type_name(key): index for index, key in enumerate(self._order)} if self._order else {}
        for sheet_name in sorted(excel.sheet_names, key=lambda key: order.get(_get_type_name(key), sys.maxsize)):
            reader = excel.sheet_reader(sheet_name)
            self._load_reader(reader, type_name=_get_type_name(sheet_name))
            self._note_issues(reader.issues, f"{file}:{sheet_name}")

    def _load_packed_file(self, file: str) -> None:
        for file in unpack_files(file, suffixes=ACCEPTABLE_FILE_SUFFIXES):
            self._load_file(file)

    def _load_json_file(self, file: str) -> None:
        with open(file) as f:
            self._add(_get_type_name(file), json.load(f))

    def _load_reader(self, reader: RowReader, type_name: str) -> None:
        schema = None
        noschema = False
        structured_column_data = _StructuredRowData(reader.header)
        for row in reader:
            if not schema and not noschema and not (schema := Schema.load_by_name(type_name, portal=self._portal)):
                noschema = True  # Create schema here just so we don't create it if there are no rows.
            structured_row = structured_column_data.create_row()
            for column_name, value in row.items():
                structured_column_data.set_value(structured_row, column_name, value, schema, reader.location)
            self._add(type_name, structured_row)

    def _add(self, type_name: str, data: Union[dict, List[dict]]) -> None:
        if self._prune:
            remove_empty_properties(data)
        if type_name in self.data:
            self.data[type_name].extend([data] if isinstance(data, dict) else data)
        else:
            self.data[type_name] = [data] if isinstance(data, dict) else data

    def _note_issues(self, issues: Optional[List[str]], source: str) -> None:
        if issues:
            if not self._issues:
                self._issues = []
            self._issues.append({source: issues})


class _StructuredRowData:

    def __init__(self, column_names: List[str]) -> None:
        self._row_template = self._parse_into_row_template(column_names)

    def create_row(self) -> dict:
        return copy.deepcopy(self._row_template)

    @staticmethod
    def set_value(row: dict, column_name: str, value: str, schema: Optional[Schema], loc: int) -> None:

        def setv(row: Union[dict, list], column_name_components: List[str], parent_array_index: int = -1) -> None:

            if not column_name_components:
                return
            if isinstance(row, list):
                if parent_array_index < 0:
                    for row_item in row:
                        setv(row_item, column_name_components)
                else:
                    setv(row[parent_array_index], column_name_components)
                return
            if not isinstance(row, dict):
                return

            column_name_component = column_name_components[0]
            array_name, array_index = _StructuredRowData._get_array_info(column_name_component)
            name = array_name if array_name else column_name_component
            if len(column_name_components) > 1:
                if not isinstance(row[name], dict) and not isinstance(row[name], list):
                    row[name] = {}
                setv(row[name], column_name_components[1:], parent_array_index=array_index)
                return

            nonlocal column_name, value, schema, loc
            if schema:
                value = schema.map_value(value, column_name, loc)
            if array_name is not None and isinstance(value, str):
                value = _split_array_string(value)
            if array_name and array_index >= 0:
                if isinstance(row[name], str):  # An array afterall e.g.: abc,abc#2
                    row[name] = _split_array_string(row[name])
                if len(row[name]) < array_index + 1:
                    row[name].extend([None] * (array_index + 1 - len(row[name])))
                row[name] = row[name][:array_index] + value + row[name][array_index + 1:]
            else:
                row[name] = value

        setv(row, _split_dotted_string(column_name))

    @staticmethod
    def _parse_into_row_template(column_names: List[str]) -> dict:

        def parse_components(column_name_components: List[str]) -> dict:
            value = parse_components(column_name_components[1:]) if len(column_name_components) > 1 else None
            column_name_component = column_name_components[0]
            array_name, array_index = _StructuredRowData._get_array_info(column_name_component)
            if array_name:
                array_length = array_index + 1 if array_index >= 0 else (0 if value is None else 1)
                # Doing it the obvious way, like in the comment right below here, we get
                # identical (shared) values; which we do not want; so do a real/deep copy.
                # return {array_name: [value] * array_length}
                return {array_name: [copy.deepcopy(value) for _ in range(array_length)]}
            return {column_name_component: value}

        structured_row_template = {}
        for column_name in column_names or []:
            if (column_name_components := _split_dotted_string(column_name)):
                merge_objects(structured_row_template, parse_components(column_name_components), True)
        return structured_row_template

    @staticmethod
    def _get_array_info(name: str) -> Tuple[Optional[str], Optional[int]]:
        if (array_indicator_position := name.rfind(ARRAY_NAME_SUFFIX_CHAR)) > 0:
            array_index = name[array_indicator_position + 1:] if array_indicator_position < len(name) - 1 else -1
            if (array_index := to_integer(array_index)) is not None:
                return name[0:array_indicator_position], array_index
        return None, None


class Schema:

    def __init__(self, schema_json: dict, portal: Optional[Portal] = None) -> None:
        self.data = schema_json
        self.name = _get_type_name(schema_json.get("title", "")) if schema_json else ""
        self._portal = portal  # Needed only to resolve linkTo references.
        self._typeinfo = self._compile_typeinfo(schema_json)

    @staticmethod
    def load_by_name(name: str, portal: Portal) -> Optional[dict]:
        return Schema(portal.get_schema(_get_type_name(name)), portal) if portal else None

    def validate(self, data: dict) -> Optional[List[str]]:
        issues = []
        for issue in SchemaValidator(self.data, format_checker=SchemaValidator.FORMAT_CHECKER).iter_errors(data):
            issues.append(issue.message)
        return issues if issues else None

    def map_value(self, value: str, column_name: str, loc: int) -> Optional[Any]:
        column_name = self._normalize_column_name(column_name)
        if (map_value := self._typeinfo.get(column_name, {}).get("map")) is None:
            map_value = self._typeinfo.get(column_name + ARRAY_NAME_SUFFIX_CHAR, {}).get("map")
        src = f"{self.name}{f'.{column_name}' if column_name else ''}{f' [{loc}]' if loc else ''}"
        return map_value(value, src) if map_value else value
        
    def _map_function(self, typeinfo: dict) -> Optional[Callable]:
        MAP_FUNCTIONS = {
            "array": self._map_function_array,
            "boolean": self._map_function_boolean,
            "enum": self._map_function_enum,
            "integer": self._map_function_integer,
            "number": self._map_function_number,
            "string": self._map_function_string
        }
        if isinstance(typeinfo, dict) and (typeinfo_type := typeinfo.get("type")) is not None:
            if isinstance(typeinfo_type, list):
                # The type specifier can actually be a list of acceptable types; for
                # example smaht-portal/schemas/mixins.json/meta_workflow_input#.value;
                # we will take the first one for which we have a mapping function.
                # TODO: Maybe more correct to get all map function and map to any for values.
                for acceptable_type in typeinfo_type:
                    if (map_function := MAP_FUNCTIONS.get(acceptable_type)) is not None:
                        break
            elif not isinstance(typeinfo_type, str):
                return None  # Invalid type specifier; ignore,
            elif isinstance(typeinfo.get("enum"), list):
                map_function = self._map_function_enum
            elif isinstance(typeinfo.get("linkTo"), str):
                map_function = self._map_function_ref
            else:
                map_function = MAP_FUNCTIONS.get(typeinfo_type)
            return map_function(typeinfo) if map_function else None
        return None

    def _map_function_array(self, typeinfo: dict) -> Callable:
        def map_array(value: str, array_type_map_function: Optional[Callable], src: Optional[str]) -> Any:
            value = _split_array_string(value)
            return [array_type_map_function(value, src) for value in value] if array_type_map_function else value
        return lambda value, src: map_array(value, self._map_function(typeinfo), src)

    def _map_function_boolean(self, typeinfo: dict) -> Callable:
        def map_boolean(value: str, src: Optional[str]) -> Any:
            return to_boolean(value, value)
        return map_boolean

    def _map_function_enum(self, typeinfo: dict) -> Callable:
        def map_enum(value: str, enum_specifiers: dict, src: Optional[str]) -> Any:
            return to_enum(value, enum_specifiers) or value
        return lambda value, src: map_enum(value, typeinfo.get("enum", []), src)

    def _map_function_integer(self, typeinfo: dict) -> Callable:
        def map_integer(value: str, src: Optional[str]) -> Any:
            return to_integer(value, value)
        return map_integer

    def _map_function_number(self, typeinfo: dict) -> Callable:
        def map_number(value: str, src: Optional[str]) -> Any:
            return to_float(value, value)
        return map_number

    def _map_function_string(self, typeinfo: dict) -> Callable:
        def map_string(value: str, src: Optional[str]) -> str:
            return value if value is not None else ""
        return map_string

    def _map_function_ref(self, typeinfo: dict) -> Callable:
        def map_ref(value: str, link_to: str, portal: Optional[Portal], src: Optional[str]) -> Any:
            nonlocal self, typeinfo
            exception = None
            if not value:
                if (column := typeinfo.get("column")) and column in self.data.get("required", []):
                    exception = f"No required reference (linkTo) value for: {link_to}"
            elif portal and not portal.ref_exists(link_to, value):
                exception = f"Cannot resolve reference (linkTo) for: {link_to}"
            if exception:
                raise Exception(exception + f"{f'/{value}' if value else ''}{f' from {src}' if src else ''}")
            return value
        return lambda value, src: map_ref(value, typeinfo.get("linkTo"), self._portal, src)

    def _compile_typeinfo(self, schema_json: dict, parent_key: Optional[str] = None) -> dict:
        """
        Given a JSON schema return a dictionary of all the property names it defines, but with
        the names of any nested properties (i.e objects within objects) flattened into a single
        property name in dot notation; and set the value of each of these flat property names
        to the type of the terminal/leaf value of the (either) top-level or nested type. N.B. We
        do NOT currently support array-of-arry or array-of-multiple-types. E.g. for this schema:

          { "properties": {
              "abc": {
                "type": "object",
                "properties": {
                  "def": { "type": "string" },
                  "ghi": {
                    "type": "object",
                    "properties": {
                      "mno": { "type": "number" }
                    }
                  }
                } },
              "stu": { "type": "array", "items": { "type": "string" } },
              "vw": {
                "type": "array",
                "items": {
                  "type": "object",
                  "properties": {
                    "xyz": { "type": "integer" }
                  } }
              } } }

        Then we will return this flat dictionary:

          { "abc.def":     { "type": "string", "map": <function:map_string> },
            "abc.ghi.mno": { "type": "number", "map": <function:map_number> },
            "stu#":        { "type": "string", "map": <function:map_string> },
            "vw#.xyz":     { "type": "integer", "map": <function:map_integer> } }
        """
        result = {}
        if (properties := schema_json.get("properties")) is None:
            if parent_key:
                if (schema_type := schema_json.get("type")) is None:
                    raise Exception(f"Array of undefined type in JSON schema NOT supported: {parent_key}")
                if schema_type == "array":
                    raise Exception(f"Array of array in JSON schema NOT supported: {parent_key}")
                result[parent_key] = {"type": schema_type, "map": self._map_function_array(schema_json)}
            return result
        for property_key, property_value in properties.items():
            if not isinstance(property_value, dict) or not property_value:
                continue  # Should not happen; every property within properties should be an object; no harm; ignore.
            key = property_key if parent_key is None else f"{parent_key}{DOTTED_NAME_DELIMITER_CHAR}{property_key}"
            if ARRAY_NAME_SUFFIX_CHAR in property_key:
                raise Exception(f"Property name with \"{ARRAY_NAME_SUFFIX_CHAR}\" in JSON schema NOT supported: {key}")
            if (property_value_type := property_value.get("type")) == "object" and "properties" in property_value:
                result.update(self._compile_typeinfo(property_value, parent_key=key))
                continue
            if property_value_type == "array":
                key += ARRAY_NAME_SUFFIX_CHAR
                if not isinstance(array_property_items := property_value.get("items"), dict):
                    if array_property_items is None or isinstance(array_property_items, list):
                        raise Exception(f"Array of undefined or multiple types in JSON schema NOT supported: {key}")
                    raise Exception(f"Invalid array type specifier in JSON schema: {key}")
                result.update(self._compile_typeinfo(array_property_items, parent_key=key))
                continue
            result[key] = {"type": property_value_type, "map": self._map_function({**property_value, "column": key})}
        return result

    @staticmethod
    def _normalize_column_name(column_name: str) -> str:
        """
        Given a string representing a flat column name, i.e possibly dot-separated name components,
        and where each component possibly ends with an array suffix (i.e. pound sign - #) followed
        by an integer, removes the integer part for each such array component; also trims names.
        For example given "abc#12. def .ghi#3" returns "abc#.def.ghi#".
        """
        return DOTTED_NAME_DELIMITER_CHAR.join([ARRAY_NAME_SUFFIX_REGEX.sub(ARRAY_NAME_SUFFIX_CHAR, value)
                                                for value in _split_dotted_string(column_name)])


class RowReader(abc.ABC):  # These readers may evenutally go into dcicutils.

    def __init__(self):
        self.header = None
        self.location = 0
        self._warning_empty_headers = False
        self._warning_extra_values = []  # Line numbers.
        self.open()

    def __iter__(self) -> Iterator:
        for row in self.rows:
            self.location += 1
            if self.is_terminating_row(row):
                break
            if len(self.header) < len(row): # Row values beyond what there are headers for are ignored.
                self._warning_extra_values.append(self.location)
            yield {column: self.cell_value(value) for column, value in zip(self.header, row)}

    def _define_header(self, header: List[Optional[Any]]) -> None:
        self.header = []
        for index, column in enumerate(header or []):
            if not (column := str(column).strip() if column is not None else ""):
                self._warning_empty_headers = True
                break  # Empty header column signals end of header.
            self.header.append(column)

    def rows(self) -> Generator[Union[List[Optional[Any]], Tuple[Optional[Any], ...]], None, None]:
        yield

    def is_terminating_row(self, row: Union[List[Optional[Any]], Tuple[Optional[Any]]]) -> bool:
        return False

    def cell_value(self, value: Optional[Any]) -> Optional[Any]:
        return str(value).strip() if value is not None else ""

    def open(self) -> None:
        pass

    @property
    def issues(self) -> Optional[List[str]]:
        issues = []
        if self._warning_empty_headers:
            issues.append("Empty header column encountered; ignoring it and all subsequent columns.")
        if self._warning_extra_values:
            issues.extend([f"Extra column values on row [{row_number}]" for row_number in self._warning_extra_values])
        return issues if issues else None


class ListReader(RowReader):

    def __init__(self, rows: List[List[Optional[Any]]]) -> None:
        self._rows = rows
        super().__init__()

    @property
    def rows(self) -> Generator[List[Optional[Any]], None, None]:
        for row in self._rows[1:]:
            yield row

    def open(self) -> None:
        if not self.header:
            self._define_header(self._rows[0] if self._rows else [])


class CsvReader(RowReader):

    def __init__(self, file: str) -> None:
        self._file = file
        self._file_handle = None
        self._rows = None
        super().__init__()

    @property
    def rows(self) -> Generator[List[Optional[Any]], None, None]:
        for row in self._rows:
            yield right_trim(row)

    def open(self) -> None:
        if self._file_handle is None:
            self._file_handle = open(self._file)
            self._rows = csv.reader(self._file_handle)
            self._define_header(right_trim(next(self._rows, [])))

    def __del__(self) -> None:
        if (file_handle := self._file_handle) is not None:
            self._file_handle = None
            file_handle.close()


class ExcelSheetReader(RowReader):

    def __init__(self, sheet_name: str, workbook: openpyxl.workbook.workbook.Workbook) -> None:
        self.sheet_name = sheet_name or "Sheet1"
        self._workbook = workbook
        self._rows = None
        super().__init__()

    @property
    def rows(self) -> Generator[Tuple[Optional[Any], ...], None, None]:
        for row in self._rows(min_row=2, values_only=True):
            yield right_trim(row)

    def is_terminating_row(self, row: Tuple[Optional[Any]]) -> bool:
        return all(cell is None for cell in row)  # Empty row signals end of data.

    def open(self) -> None:
        if not self._rows:
            self._rows = self._workbook[self.sheet_name].iter_rows
            self._define_header(right_trim(next(self._rows(min_row=1, max_row=1, values_only=True), [])))


class Excel:

    def __init__(self, file: str) -> None:
        self._file = file
        self._workbook = None
        self.sheet_names = None
        self.open()

    def sheet_reader(self, sheet_name: str) -> ExcelSheetReader:
        return ExcelSheetReader(sheet_name=sheet_name, workbook=self._workbook)

    def open(self) -> None:
        if self._workbook is None:
            self._workbook = openpyxl.load_workbook(self._file, data_only=True)
            self.sheet_names = self._workbook.sheetnames or []

    def __del__(self) -> None:
        if (workbook := self._workbook) is not None:
            self._workbook = None
            workbook.close()


class Portal:

    def __init__(self, portal: PortalAny, data: Optional[dict] = None, schemas: Optional[List[dict]] = None) -> None:
        self.vapp = portal.vapp if isinstance(portal, Portal) else portal
        self._data = data  # Data set being loaded (e.g. by StructuredDataSet).
        self._schemas = schemas  # Explicitly specified known schemas.

    @lru_cache(maxsize=256)
    def get_schema(self, schema_name: str) -> Optional[dict]:
        return (next((schema for schema in self._schemas or []
                     if _get_type_name(schema.get("title")) == _get_type_name(schema_name)), None) or
                get_schema(schema_name, portal_vapp=self.vapp))

    @lru_cache(maxsize=256)
    def get_metadata(self, object_name: str) -> Optional[dict]:
        try:
            return get_metadata(object_name, vapp=self.vapp)
        except Exception:
            return None

    def ref_exists(self, type_name: str, value: str) -> bool:
        if self._data and (items := self._data.get(type_name)) and (schema := self.get_schema(type_name)):
            id_properties = set(schema.get("identifyingProperties", [])) | {"identifier", "uuid"}
            for item in items:
                for id_property in id_properties:
                    if (id_value := item.get(id_property)) is not None:
                        if isinstance(id_value, list) and value in id_value or id_value == value:
                            return True
        return self.get_metadata(f"/{type_name}/{value}") is not None

    @staticmethod
    def create(portal: Optional[PortalAny] = None,
               data: Optional[dict] = None, schemas: Optional[List[dict]] = None) -> Optional[Portal]:
        if isinstance(portal, Portal):
            if data is not None:
                portal._data = data
            if schemas is not None:
                portal._schemas = schemas
            return portal
        return Portal(portal, data=data, schemas=schemas) if portal else None

    @staticmethod
    def create_for_testing(ini_file: Optional[str] = None, schemas: Optional[List[dict]] = None) -> Portal:
        if isinstance(ini_file, str):
            return Portal(create_testapp(ini_file), schemas=schemas)
        minimal_ini_for_unit_testing = "[app:app]\nuse = egg:encoded\nsqlalchemy.url = postgresql://dummy\n"
        with temporary_file(content=minimal_ini_for_unit_testing, suffix=".ini") as ini_file:
            return Portal(create_testapp(ini_file), schemas=schemas)


def _split_dotted_string(value: str):
    return split_string(value, DOTTED_NAME_DELIMITER_CHAR)


def _split_array_string(value: str):
    return split_string(value, ARRAY_VALUE_DELIMITER_CHAR, ARRAY_VALUE_DELIMITER_ESCAPE_CHAR)


def _get_type_name(value: str) -> str:  # File or other name.
    name = os.path.basename(value).replace(" ", "") if isinstance(value, str) else ""
    return to_camel_case(name[0:dot] if (dot := name.rfind(".")) > 0 else name)
