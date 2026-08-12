from __future__ import annotations

import argparse
import dataclasses
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import googleapiclient
import googleapiclient.discovery
import openpyxl
import structlog
from dcicutils import ff_utils
from dcicutils.creds_utils import SMaHTKeyManager
from dcicutils.misc_utils import to_camel_case, to_snake_case
from dcicutils import schema_utils
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from snovault.schema_views import SubmissionSchemaConstants

from encoded.item_utils.constants import item as item_constants
from encoded.item_utils.utils import RequestHandler
from encoded.project.loadxl import ITEM_INDEX_ORDER

log = structlog.getLogger(__name__)

GOOGLE_SHEET_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
GOOGLE_CREDENTIALS_LOCATION = "~/google_sheets_creds.json"
GOOGLE_CREDENTIALS_PATH = Path.expanduser(Path(GOOGLE_CREDENTIALS_LOCATION))
GOOGLE_TOKEN_PATH = Path.expanduser(Path("~/google_sheets_token.json"))

"""
Google Information as of 2024-06-10
===================================

Google Sheets API Reference:
    * https://developers.google.com/sheets/api/reference/rest/v4

To generate credentials for Google Sheets API, see instructions under
the Desktop app section here:
    * https://developers.google.com/workspace/guides/create-credentials

Overview:
1. Go to Google Cloud Console: https://console.cloud.google.com/
2. Create a new project
3. Enable Google Sheets API
4. Create credentials for the project
5. Download the credentials as JSON
6. Save the credentials as `google_sheets_creds.json` in the home directory
7. Under OAuth consent screen, add your email under Test users
8. Run the script to generate the token
9. Token will be saved as `google_sheets_token.json` in the home directory

If token expires, delete the token file and run the script again to generate
a new token.
"""

ITEM_SPREADSHEET_SUFFIX = "_submission.xlsx"
WORKBOOK_FILENAME = "submission_workbook.xlsx"

EXAMPLE_FILE_UUIDS=["4e142999-5d48-4dcd-b7d6-558e5960e69b",
                    "d4020a63-338c-4103-8461-417d09df5cbd",
                    "f382f84c-f590-49f2-9f62-e852e2c30647"
                    ]

POPULATE_ORDER = [
    "VariantCalls",
    "AlignedReads",
    "UnalignedReads",
    "FileSet",
    "Software",
    "Library",
    "Sequencing",
    "LibraryPreparation",
    "Analyte",
    "AnalytePreparation",
    "PreparationKit",
    "Treatment",
    "CellSample",
    "CellCultureSample",
    "CellCultureMixture",
    "CellCulture",
    "CellLine",
    "TissueSample",
    "Tissue",
    "Donor"
]

FONT = "Arial"
FONT_SIZE = 10

TPC_SUBMISSION_ITEMS = [
    "Donor",
    "Demographic",
    "MedicalHistory",
    "Diagnosis",
    "Exposure",
    "FamilyHistory",
    "MedicalTreatment",
    "DeathCircumstances",
    "TissueCollection",
    "Tissue",
    "TissueSample",
    "NonBrainPathologyReport",
    "BrainPathologyReport",
    "HistologyImage"
]

GCC_SUBMISSION_ITEMS = [
    "Donor",
    "Tissue",
    "TissueSample",
    "CellLine",
    "CellCulture",
    "CellCultureMixture",
    "CellCultureSample",
    "CellSample",
    "Analyte",
    "AnalytePreparation",
    "PreparationKit",
    "Treatment",
    "Library",
    "LibraryPreparation",
    "Sequencing",
    "FileSet",
    "UnalignedReads",
    "AlignedReads",
    "VariantCalls",
    "SupplementaryFile",
    "Software"
]

DSA_SUBMISSION_ITEMS = [
    "DonorSpecificAssembly",
    "SupplementaryFile",
    "Software"
]

EQM_TAB_NAMES = {
    "dsa": "DSA_ExternalQualityMetric",
    "duplexseq": "DuplexSeq_ExternalQualityMetric"
}
MULTI_TYPE_ITEMS = [
    'CellCultureMixture'
]

PSEUDO_PROPERTIES = {
    "FileSet": {
        "expected_file_count": {
            "title": "Expected File Count",
            "description": "Number of files expected to be within the file set",
            "type": "integer",
            "is_required": "True"
        }
    }
}


@dataclass(frozen=True)
class SheetsClient:
    client: googleapiclient.discovery.Resource
    sheet_id: str

    def get_worksheets(self) -> List[Dict[str, Any]]:
        """Get the worksheets from Google Sheets."""
        sheet = self.client.get(spreadsheetId=self.sheet_id).execute()
        return sheet["sheets"]

    def submit_requests(self, requests: List[Dict[str, Any]]) -> None:
        """Submit requests to Google Sheets."""
        self.client.batchUpdate(
            spreadsheetId=self.sheet_id,
            body={"requests": requests},
        ).execute()


def update_google_sheets(
    sheets_client: SheetsClient,
    request_handler: RequestHandler,
    gcc: bool = False,
    tpc: bool = False,
    items: List[str] = None,
    eqm: Union[str, None] = None,
    example: bool = False,
) -> None:
    """Update Google Sheets with the latest submission schemas."""
    spreadsheets = get_spreadsheets(request_handler, gcc=gcc, tpc=tpc, items=items, eqm=eqm, example=example)
    log.info("Clearing existing Google sheets.")
    delete_existing_sheets(sheets_client)
    log.info("Updating Google sheets with tabs.")
    update_or_add_spreadsheets(sheets_client, spreadsheets)
    log.info("Writing properties to Google sheets.")
    write_values_to_sheets(sheets_client, spreadsheets)
    log.info("Formatting columns in Google sheets.")
    format_column_widths(sheets_client, spreadsheets)
    log.info("Google sheets updated.")


def get_spreadsheets(
        request_handler: RequestHandler,
        gcc: bool = False,
        tpc: bool = False,
        items: List[str] = None,
        eqm: Union[str, None] = None,
        example: bool = False,
    ) -> List[Spreadsheet]:
    submission_schemas = get_all_submission_schemas(request_handler)
    ordered_submission_schemas = get_ordered_submission_schemas(submission_schemas, gcc=gcc, tpc=tpc)
    if example:
        example_fields = get_example_fields(EXAMPLE_FILE_UUIDS,GCC_SUBMISSION_ITEMS)
        submission_schemas = get_ordered_submission_schemas(submission_schemas,order=POPULATE_ORDER)
        spreadsheets = []
        for item, submission_schema in submission_schemas.items():
            unlinked_spreadsheet = get_example_spreadsheet(
                item,request_handler,example_fields,submission_schema
            )
            spreadsheet,example_fields = get_linked_spreadsheet(
                request_handler,unlinked_spreadsheet,example_fields
            )
            spreadsheets.append(spreadsheet)
        if gcc:
            order = GCC_SUBMISSION_ITEMS
        elif items:
            order = items
        return reorder_spreadsheets(spreadsheets,order)
    if items:
        spreadsheets = [
            get_spreadsheet(item, submission_schema)
            for item, submission_schema in ordered_submission_schemas.items()
            if item in items
        ]
    else:
        spreadsheets = [
            get_spreadsheet(item, submission_schema)
            for item, submission_schema in ordered_submission_schemas.items()
        ]
    if eqm:
        eqm_schema ={
            'schema': get_submission_schema('ExternalQualityMetric', request_handler)
        }
        spreadsheets.append(get_eqm_spreadsheet(eqm, eqm_schema, request_handler))
    return spreadsheets


def reorder_spreadsheets(spreadsheets: List[Spreadsheet], order: List[str]):
    """Reorder spreadsheets in a specific order."""
    new_spreadsheets = []
    for item in order:
        for spreadsheet in spreadsheets:
            if spreadsheet.item == item:
                new_spreadsheets.append(spreadsheet)
    return new_spreadsheets


def delete_existing_sheets(sheets_client: SheetsClient) -> None:
    """Delete existing sheets from Google Sheets."""
    requests = []
    for sheet in sheets_client.get_worksheets():
        sheet_id = get_worksheet_id(sheet)
        if sheet_id == 0:
            requests.append(get_clear_values_request(sheet_id))
        else:
            requests.append(get_delete_sheet_request(sheet_id))
    if requests:
        sheets_client.submit_requests(requests)


def get_worksheet_id(sheet: Dict[str, Any]) -> int:
    """Get the worksheet ID."""
    return sheet["properties"]["sheetId"]


def get_clear_values_request(sheet_id: int) -> Dict[str, Any]:
    """Get request to clear values from the sheet."""
    return {"deleteRange": {"range": {"sheetId": sheet_id}, "shiftDimension": "ROWS"}}


def get_delete_sheet_request(sheet_id: int) -> Dict[str, Any]:
    """Get request to delete the sheet."""
    return {"deleteSheet": {"sheetId": sheet_id}}


def update_or_add_spreadsheets(
    sheets_client: SheetsClient, spreadsheets: List[Spreadsheet]
) -> None:
    """Update or add spreadsheets to Google Sheets."""
    requests = []
    for idx, spreadsheet in enumerate(spreadsheets):
        if idx == 0:
            requests.append(get_update_sheet_title_request(spreadsheet, idx))
        else:
            requests.append(get_add_sheet_request(spreadsheet, idx))
    if requests:
        sheets_client.submit_requests(requests)


def get_update_sheet_title_request(
    spreadsheet: Spreadsheet, sheet_id: int
) -> Dict[str, Any]:
    """Get request to update the sheet title."""
    return {
        "updateSheetProperties": {
            "properties": {
                "sheetId": sheet_id,
                "title": spreadsheet.item,
                "gridProperties": {
                    "columnCount": get_max_column_count(spreadsheet.properties)
                },
            },
            "fields": "title",
        }
    }


def get_max_column_count(properties: List[Property]) -> int:
    """Get the maximum column count."""
    return len(properties) if len(properties) > 15 else 15


def get_add_sheet_request(spreadsheet: Spreadsheet, sheet_id: int) -> Dict[str, Any]:
    """Get request to add a new sheet."""
    return {
        "addSheet": {
            "properties": {
                "title": spreadsheet.item,
                "sheetId": sheet_id,
                "gridProperties": {
                    "columnCount": get_max_column_count(spreadsheet.properties)
                },
            },
        }
    }


def write_values_to_sheets(
    sheets_client: SheetsClient,
    spreadsheets: List[Spreadsheet]
) -> None:
    """Write values to the Google Sheets."""
    requests = []
    for idx, spreadsheet in enumerate(spreadsheets):
        requests.append(get_update_cells_request(spreadsheet, idx))
    if requests:
        sheets_client.submit_requests(requests)


def get_update_cells_request(spreadsheet: Spreadsheet, sheet_id: int) -> Dict[str, Any]:
    """Get request to update cells with properties."""
    if spreadsheet.examples:
        values = get_example_values(spreadsheet)
    else: 
        values = get_values(spreadsheet)
    
    return {
        "updateCells": {
            "rows": values,
            "fields": "*",
            "start": {"sheetId": sheet_id, "rowIndex": 0, "columnIndex": 0},
        }
    }


def get_values(spreadsheet: Spreadsheet) -> List[Dict[str, Any]]:
    """Get values for the spreadsheet."""
    ordered_properties = get_ordered_properties(spreadsheet.properties)
    return [{"values": [get_cell_value(property_) for property_ in ordered_properties]}]


def get_cell_value(property_: Property) -> Dict[str, Any]:
    """Get the cell value."""
    return {
        "userEnteredValue": {"stringValue": property_.name},
        "userEnteredFormat": {
            "textFormat": get_text_format(property_),
        },
        "note": get_comment_text(property_),
    }


def get_example_cell_value(value: Union[str, List]) -> Dict[str, Any]:
    """Get the example cell value."""
    if is_list(value):
        value=get_example_list(value)
    return {
        "userEnteredValue": {"stringValue": str(value)},
        "userEnteredFormat": {
            "textFormat": {"fontFamily": FONT, "fontSize": FONT_SIZE}
        }
    }


def get_empty_cell():
    """Get empty cell"""
    return {
        "userEnteredValue": {},
        "userEnteredFormat": {}
    }


def get_example_values(spreadsheet: Spreadsheet) -> List[Dict[str, Any]]:
    """Get example values for the Google spreadsheet."""
    ordered_properties = get_ordered_properties(spreadsheet.properties)
    values = [{"values": [get_cell_value(property_) for property_ in ordered_properties]}]
    for idx, example in enumerate(spreadsheet.examples):
        row_values = {"values":[]}
        for property_ in ordered_properties:
            if property_.nested:
                parent_property, nested_property, n_index = extract_nested_property_names(property_.name)
                value = example[parent_property][n_index][nested_property]
                row_values["values"].append(get_example_cell_value(value))
            elif property_.name in example:
                value = example[property_.name]
                row_values["values"].append(get_example_cell_value(value))
            else:
                row_values["values"].append(get_empty_cell())
        values.append(row_values)
    return values


def get_text_format(property_: Property) -> Dict[str, Any]:
    """Get the text format."""
    text_format = {"fontFamily": FONT, "fontSize": FONT_SIZE}
    if property_.required:
        text_format["bold"] = True
    if property_.link:
        text_format["italic"] = True
    return text_format


def format_column_widths(
    sheets_client: SheetsClient, spreadsheets: List[Spreadsheet]
) -> None:
    """Format column widths in the Google Sheets."""
    requests = []
    column_width_multiplier = 7  # 7 pixels per character seemed to work well
    for idx, spreadsheet in enumerate(spreadsheets):
        for index, property_ in enumerate(spreadsheet.properties):
            width = len(property_.name) * column_width_multiplier
            requests.append(get_format_column_request(idx, index, width))
    if requests:
        sheets_client.submit_requests(requests)


def get_format_column_request(
    sheet_id: int, column_index: int, width: int
) -> Dict[str, Any]:
    """Get request to format an individual column."""
    minimum_width = 120  # Keep a minimum width of 120 pixels for the columns
    width = width if width > minimum_width else minimum_width
    return {
        "updateDimensionProperties": {
            "range": {
                "sheetId": sheet_id,
                "dimension": "COLUMNS",
                "startIndex": column_index,
                "endIndex": column_index + 1,
            },
            "properties": {
                "pixelSize": width,
            },
            "fields": "pixelSize",
        }
    }


def write_all_spreadsheets(
    output: Path,
    request_handler: RequestHandler,
    workbook: bool = False,
    separate_comments: bool = False
) -> None:
    """Write all submission spreadsheets"""
    submission_schemas = get_all_submission_schemas(request_handler)
    log.info(f"Writing submission spreadsheets to: {output}")
    if workbook:
        write_workbook(output, submission_schemas, request_handler, separate_comments=separate_comments)
    else:
        write_spreadsheets(
            output, submission_schemas, request_handler, separate_comments=separate_comments
        )


def get_all_submission_schemas(
    request_handler: RequestHandler,
) -> Dict[str, Dict[str, Any]]:
    """Get all submission schemas"""
    return request_handler.get_item(SubmissionSchemaConstants.ENDPOINT)


def write_item_spreadsheets(
    output: Path,
    items: List[str],
    request_handler: RequestHandler,
    workbook: bool = False,
    tpc: bool = False,
    gcc: bool = False,
    eqm: Union[str, None] = None,
    separate_comments: bool = False,
    example: bool = False
) -> None:
    """Write submission spreadsheets for specified items"""
    submission_schemas = get_submission_schemas(items, request_handler)
    eqm_schema = None
    if eqm:
        eqm_schema = {
            'schema': get_submission_schema('ExternalQualityMetric', request_handler)
        }
    if not submission_schemas:
        log.error("No submission schemas found for given items. Exiting...")
        return
    if example:
        log.info(
            f"Writing example submission spreadsheets to {output} for items:"
            f" {submission_schemas.keys()}"
        )
    else:
        log.info(
            f"Writing submission spreadsheets to {output} for items:"
            f" {submission_schemas.keys()}"
        )
    if workbook:
        write_workbook(output, submission_schemas, request_handler, separate_comments=separate_comments, tpc=tpc, gcc=gcc, eqm=eqm, eqm_schema=eqm_schema, example=example)
    else:
        write_spreadsheets(
            output, submission_schemas, request_handler, separate_comments=separate_comments, eqm=eqm, eqm_schema=eqm_schema, example=example
        )


def get_submission_schemas(
    items: List[str], request_handler: RequestHandler
) -> Dict[str, Dict[str, Any]]:
    """Get submission schemas for items."""
    submission_schemas = {
        to_camel_case(item): get_submission_schema(item, request_handler)
        for item in items
    }
    return {key: value for key, value in submission_schemas.items() if value}


def get_submission_schema(item: str, request_handler: RequestHandler) -> Dict[str, Any]:
    """Get the submission schema for the item."""
    try:
        return request_handler.get_item(get_submission_schema_endpoint(item))
    except Exception as e:
        log.error(f"Error getting submission schema for {item}: {e}")
        return {}


def get_submission_schema_endpoint(item: str) -> Dict[str, Any]:
    """Get the submission schema for the item"""
    return f"{SubmissionSchemaConstants.ENDPOINT}{to_snake_case(item)}.json"


def write_workbook(
    output: Path,
    submission_schemas: Dict[str, Any],
    request_handler: RequestHandler,
    tpc: bool = False,
    gcc: bool = False,
    separate_comments: bool = False,
    eqm: Union[str, None] = None,
    eqm_schema: Union[Dict[str, Any], None] = None,
    example: bool = False
) -> None:
    """Write a single workbook containing all submission spreadsheets."""
    workbook = openpyxl.Workbook()
    ordered_submission_schemas = get_ordered_submission_schemas(submission_schemas,tpc=tpc,gcc=gcc)
    write_workbook_sheets(
        workbook, ordered_submission_schemas, request_handler, separate_comments=separate_comments, eqm=eqm, eqm_schema=eqm_schema, example=example
    )
    file_path = Path(output, WORKBOOK_FILENAME)
    save_workbook(workbook, file_path)
    if example:
        log.info(f"Example workbook written to: {file_path}")
    else:
        log.info(f"Workbook written to: {file_path}")


def get_ordered_submission_schemas(
    submission_schemas: Dict[str, Any],
    tpc: bool = False,
    gcc: bool = False,
    order: List[str] = None
) -> Dict[str, Dict[str, Any]]:
    """Order submission schemas."""
    result = {}
    if order:
        item_order = order
    elif tpc:
        item_order = TPC_SUBMISSION_ITEMS
    elif gcc:
        item_order = GCC_SUBMISSION_ITEMS
    else:
        item_order = [to_camel_case(item) for item in ITEM_INDEX_ORDER]
    for item in item_order:
        if item in submission_schemas:
            result[item] = submission_schemas[item]
    return result


@dataclass
class ExampleFields:
    """Data struct for keeping track of linked items in example spreadsheets.
    
    Seed File is currently set with EXAMPLE_FILE_UUIDs."""
    seed_files: List[str]
    fields: Dict[str,List[Union[str,None]]]


def get_example_fields(seed_files: List[str],items = List[str]) -> ExampleFields:
    """Get linked id information for item."""
    fields = {}
    for item in items:
        fields[item] = []
    return ExampleFields(
        seed_files=seed_files,
        fields=fields,
    )


def write_workbook_sheets(
    workbook: openpyxl.Workbook,
    submission_schemas: Dict[str, Dict[str, Any]],
    request_handler: RequestHandler,
    separate_comments: bool = False,
    eqm: Union[str, None] = None,
    eqm_schema: Union[Dict[str, Any], None] = None,
    example: bool = False
) -> None:
    """Write workbook sheets for given schemas."""
    if example:
        spreadsheets = []
        example_fields = get_example_fields(EXAMPLE_FILE_UUIDS,GCC_SUBMISSION_ITEMS)
        submission_schemas = get_ordered_submission_schemas(submission_schemas,order=POPULATE_ORDER)
        for index, (item, submission_schema) in enumerate(submission_schemas.items()):
            unlinked_spreadsheet = get_example_spreadsheet(
                item, request_handler, example_fields, submission_schema
            )
            spreadsheet, example_fields = get_linked_spreadsheet(
                request_handler, unlinked_spreadsheet, example_fields
            )
            spreadsheets.append(spreadsheet)
        ordered_spreadsheets = reorder_spreadsheets(spreadsheets, GCC_SUBMISSION_ITEMS)
    else:
        ordered_spreadsheets = []
        for index, (item, submission_schema) in enumerate(submission_schemas.items()):
            spreadsheet = get_spreadsheet(item, submission_schema)
            ordered_spreadsheets.append(spreadsheet)
    for index, spreadsheet in enumerate(ordered_spreadsheets):
        if index == 0:
            worksheet = workbook.active
            set_sheet_name(worksheet, spreadsheet)
            write_properties(worksheet, spreadsheet.properties, separate_comments, examples=spreadsheet.examples)
        else:
            worksheet = workbook.create_sheet(title=spreadsheet.item)
            write_properties(worksheet, spreadsheet.properties, separate_comments)
    if eqm:
        spreadsheet = get_eqm_spreadsheet(eqm, eqm_schema, request_handler)
        worksheet = workbook.create_sheet(title=spreadsheet.item)
        write_properties(worksheet, spreadsheet.properties, separate_comments)


def write_spreadsheets(
    output: Path,
    submission_schemas: Dict[str, Any],
    request_handler: RequestHandler,
    separate_comments: bool = False,
    eqm: Union[str, None] = None,
    eqm_schema: Union[Dict[str, Any], None] = None,
    example: bool = False
) -> None:
    """Write submission spreadsheets."""
    if example:
        example_fields = get_example_fields(EXAMPLE_FILE_UUIDS,GCC_SUBMISSION_ITEMS)
        submission_schemas = get_ordered_submission_schemas(submission_schemas,order=POPULATE_ORDER)
        for item, submission_schema in submission_schemas.items():
            unlinked_spreadsheet = get_example_spreadsheet(
                item,request_handler,example_fields,submission_schema
            )
            spreadsheet,example_fields = get_linked_spreadsheet(
                request_handler,unlinked_spreadsheet,example_fields
            )
            write_spreadsheet(output, spreadsheet, separate_comments, example=example)
    else:
        for item, submission_schema in submission_schemas.items():
            spreadsheet = get_spreadsheet(item, submission_schema)
            write_spreadsheet(output, spreadsheet, separate_comments)
        if eqm:
            spreadsheet = get_eqm_spreadsheet(eqm, eqm_schema, request_handler)
            write_spreadsheet(output, spreadsheet, separate_comments)


@dataclass(frozen=True)
class Property:
    """Struct to hold property info required for spreadsheet.

    Arrays of objects with nested properties are handled by making new
    `name#N.child` Property instances for each slot (currently relevant for
    CellCultureMixture and PathologyReport items, and configurable via
    `get_array_slot_count`). Plain (non-array) nested objects are flattened to
    `name.child` Property instances (see `get_flattened_object_properties`).
    Arrays of strings are handled by bringing select info to top level.
    """

    name: str
    item: str = ""
    description: str = ""
    value_type: str = ""
    required: bool = False
    link: bool = False
    enum: Optional[List[str]] = None
    array_subtype: str = ""
    pattern: str = ""
    comment: str = ""
    examples: Optional[List[str]] = None
    format_: str = ""
    requires: Optional[List[str]] = None
    exclusive_requirements: Optional[List[str]] = None
    nested: bool = False
    search: str = ""
    allow_commas: Optional[bool] = False
    allow_multiplier_suffix: Optional[bool] = False
    search: Optional[str] = ""


@dataclass(frozen=True)
class Spreadsheet:
    item: str
    properties: List[Property]
    examples: Optional[List[Dict[str,Any]]] = None


def get_spreadsheet(item: str, submission_schema: Dict[str, Any]) -> Spreadsheet:
    """Get spreadsheet information for item."""
    properties = get_properties(item, submission_schema)
    return Spreadsheet(
        item=item,
        properties=properties,
    )


def get_eqm_spreadsheet(eqm: str, eqm_schema: Dict[str, Any], request_handler: RequestHandler):
    """Get spreadsheet information for ExternalQualityMetric item."""
    item = EQM_TAB_NAMES[eqm]
    result = get_eqm_mapping(request_handler)
    if item in result['sheet_mappings']:
        column_mapping = result['sheet_mappings'][item]
        mapping = result['column_mappings'][column_mapping]
        properties = get_eqm_properties(item, eqm, eqm_schema, mapping)
        return Spreadsheet(
            item=item,
            properties=properties,
        )
    else:
        log.error("No ExternalQualityMetric found for given `eqm` value. Exiting...")
        return


def get_eqm_mapping(request_handler: RequestHandler):
    """Get JSON mapping config item from portal query."""
    search = "search/?type=GenericQcConfig&tags=external_quality_metrics"
    result = ff_utils.search_metadata(search, key=request_handler.auth_key)
    return result[0]['body']


def get_example_spreadsheet(
        item: str,
        request_handler: RequestHandler,
        example_fields: ExampleFields,
        submission_schema: Dict[str,Any]
    ) -> Spreadsheet:
    """Get example property values of spreadsheet information for item."""
    starting = ['AlignedReads'] # Currently just aligned reads
    #starting = ['AlignedReads','VariantCalls']
    if item in starting:
        example = get_submission_examples(request_handler,example_fields,seed=True)
    else:
        example = get_submission_examples(request_handler,example_fields,item_type=item)
    properties = get_properties(item, submission_schema)
    return Spreadsheet(
        item=item,
        properties=properties,
        examples=example
    )


def get_submission_examples(
    request_handler: RequestHandler,
    example_fields: ExampleFields,
    item_type: str = None,
    seed: bool = False
    ):
    """Get examples of property values for items."""
    items = example_fields.seed_files if seed else example_fields.fields[item_type]
    return [request_handler.get_item(obj_id) for obj_id in items]


def get_linked_spreadsheet(
    request_handler: RequestHandler,
    spreadsheet: Spreadsheet,
    example_fields: ExampleFields
):
    """Get spreadsheet with links filled out with submitted_id or identifier."""
    links = get_all_links(spreadsheet)
    nested_links = get_nested_links(spreadsheet)
    for link in links:
        for idx, example in enumerate(spreadsheet.examples):
            if link in nested_links:   
                values = get_nested_example(example,link)
                parent_property, nested_property, n_index = extract_nested_property_names(link)
                id_values, example_fields = get_id_list(request_handler,values,example_fields)
                spreadsheet.examples[idx][parent_property][n_index][nested_property] = " | ".join(id_values)
            elif link in example:
                values = example[link]
                id_values, example_fields = get_id_list(request_handler,values,example_fields)
                spreadsheet.examples[idx][link] = " | ".join(id_values)
    return spreadsheet, example_fields


def get_id_list(
    request_handler: RequestHandler,
    values: Union[str,List[str]],
    example_fields: ExampleFields
    ):
    """Return list of submitted_id or identifier values from @ids."""
    id_values = []
    if type(values) is list:
        for value in values:
            item = request_handler.get_item(value)
            example_fields = update_example_fields(item,example_fields)
            id_values.append(get_linked_item_id(item))
    else:
        item = request_handler.get_item(values)
        example_fields = update_example_fields(item,example_fields)
        id_values.append(get_linked_item_id(item))
    return id_values, example_fields


def update_example_fields(item: Dict[str,Any],example_fields: ExampleFields):
    """Update value of example field for linked submitted item.
    
    For item types with multiple submitted items in the type list, get the type to use from MULTI_TYPE_ITEMS. Currently, the only item that matches multiple keys is CellCultureMixture but may need to update later.
    """
    key = [value for value in item['@type'] if value in example_fields.fields.keys()]
    if len(key)>1:
        overlap = [match for match in set(MULTI_TYPE_ITEMS) & set(key)]
        key = overlap
    if key:
        atid = item.get("@id","")
        if atid not in example_fields.fields[key[0]]:
            example_fields.fields[key[0]].append(atid) 
    return example_fields


def get_linked_item_id(response: Dict[str,Any]):
    """Get either submitted_id or identifier for item."""
    submitted_id = response.get("submitted_id","")
    if submitted_id:
        return submitted_id
    else:
        return response.get("identifier","")
        

def get_all_links(spreadsheet: Spreadsheet):
    """Get all links from properties."""
    links = get_required_links(spreadsheet.properties) + get_non_required_links(spreadsheet.properties)
    return [link.name for link in links]


def get_nested_links(spreadsheet: Spreadsheet) -> List[Property]:
    """Get links that are nested within properties."""
    return [
            property_.name
            for property_ in spreadsheet.properties
            if property_.nested and property_.link
        ]


def get_properties(item: str, submission_schema: Dict[str, Any]) -> List[Property]:
    """Get property information from the submission schema.
     
     Add any special pseudo-properties that are item-specific.
    """
    properties = schema_utils.get_properties(submission_schema)
    if item in PSEUDO_PROPERTIES.keys():
        properties = {**properties,**PSEUDO_PROPERTIES[item]}
    property_list = []
    for key, value in properties.items():
        property_list += get_nested_properties(item, key, value)
    return property_list


def get_eqm_properties(
        item: str,
        eqm: Dict[str, Any],
        eqm_schema: Dict[str, Any],
        mapping: Dict[str, Any]
    ):
    """Format property information from ExternalQualityMetric template.
    
    Grabs normal properties from schema and then formats qc_values.key and qc_values.tooltip to be description and submissionComment.
    """
    properties = schema_utils.get_properties(eqm_schema['schema'])
    primary_properties = {key: value for key, value in properties.items() if key != "qc_values"}
    secondary_properties = get_eqm_qc_values(item, mapping)
    all_properties = {**primary_properties, **secondary_properties}
    property_list = []
    for key, value in all_properties.items():
        property_list.append(get_property(item, key, value))
    return property_list
    

def get_eqm_qc_values(item: str, qc_values: Dict[str, Any]):
    """Get qc_values format to match schema properties.
    
    If `tooltip` is present, add as a `submissionComment.
    """
    qc_values_properties = {}
    for metric, values in qc_values.items():
        qc_values_properties[metric] = {
            "description": values['qc_values#.key'],
            "type": values['qc_values#.value'].split(':')[1].split('}')[0]
        }
        if "qc_values#.tooltip" in values:
            qc_values_properties[metric]["submissionComment"] = values['qc_values#.tooltip']
    return qc_values_properties


def get_property(
    item: str,
    property_name: str,
    property_schema: Dict[str, Any],
    is_nested: bool = False,
    required_override: Optional[bool] = None,
) -> Property:
    """Get property information.

    `required_override`, if not None, takes precedence over `is_required()`.
    Needed for schema sources (e.g. raw profile schemas, or nested object/array
    items) that don't carry the submission-schema endpoint's computed
    `is_required` annotation but do have their own `required` list to derive it
    from instead.
    """
    required = is_required(property_schema) if required_override is None else required_override
    return Property(
        name=property_name,
        item=item,
        description=schema_utils.get_description(property_schema),
        value_type=schema_utils.get_schema_type(property_schema),
        required=required,
        link=is_link(property_schema),
        enum=get_enum(property_schema),
        array_subtype=get_array_subtype(property_schema),
        pattern=schema_utils.get_pattern(property_schema),
        comment=schema_utils.get_submission_comment(property_schema),
        examples=get_examples(property_schema),
        format_=schema_utils.get_format(property_schema),
        requires=get_corequirements(property_schema),
        exclusive_requirements=get_exclusive_requirements(property_schema),
        nested=is_nested,
        allow_commas=is_allow_commas(property_schema),
        allow_multiplier_suffix=is_allow_multiplier_suffix(property_schema),
        search=get_search_url(property_schema)
    )


DEFAULT_ARRAY_SLOT_COUNT = 2


def get_nested_properties(
    item: str,
    property_name: str,
    property_schema: Dict[str, Any],
    array_slot_counts: Optional[Dict[str, int]] = None,
    default_array_slot_count: int = DEFAULT_ARRAY_SLOT_COUNT,
    required_override: Optional[bool] = None,
) -> List[Property]:
    """Get nested property information for a property.

    Handles, in order: arrays of objects (flattened to `name#N.child` columns),
    plain nested objects (flattened to `name.child` columns), and otherwise
    falls back to a single scalar/link property.

    `array_slot_counts`/`default_array_slot_count` configure how many `#N` slots
    to generate for array-of-object properties; see `get_array_slot_count`.
    `required_override` is only consulted for the scalar fallback case, since the
    array/object branches derive their children's required-ness from the nested
    schema's own `required` list instead.
    """
    if object_array := get_array_object_properties(property_schema):
        count = get_array_slot_count(
            item, property_name, object_array, array_slot_counts, default_array_slot_count
        )
        required_keys = get_object_required(property_schema.get("items", {}))
        return get_nested_property(item, property_name, object_array, count=count, required_keys=required_keys)
    if object_properties := get_plain_object_properties(property_schema):
        required_keys = get_object_required(property_schema)
        return get_flattened_object_properties(item, property_name, object_properties, required_keys)
    return [get_property(item, property_name, property_schema, required_override=required_override)]


def get_array_slot_count(
    item: str,
    property_name: str,
    property_schema: Dict[str, Any],
    array_slot_counts: Optional[Dict[str, int]],
    default_count: int = DEFAULT_ARRAY_SLOT_COUNT,
) -> int:
    """Get the number of `#N` slot columns to generate for an array-of-objects property.

    Checks an explicit per-property override first (`array_slot_counts`, keyed by
    property name), then falls back to the legacy per-item special-casing for
    PathologyReport items (kept for backward compatibility with existing
    templates), then `default_count`.
    """
    if array_slot_counts and property_name in array_slot_counts:
        return array_slot_counts[property_name]
    if item == "NonBrainPathologyReport":
        # target tissues stays with count 2 as that's usually how many there are
        # non-target tissues and pathologic findings have as many as there are enums
        if property_name == 'non_target_tissues':
            return len(property_schema['non_target_tissue_subtype']['enum'])
        elif property_name == 'pathologic_findings':
            return len(property_schema['finding_type']['enum'])
    elif item == "BrainPathologyReport":
        # brain_subregions has as many as there are enums
        if property_name == 'brain_subregions':
            return len(property_schema['subregion']['enum'])
    return default_count


def get_nested_property(
    item: str,
    property_name: str,
    property_schema: Dict[str, Any],
    count: int = DEFAULT_ARRAY_SLOT_COUNT,
    required_keys: Optional[List[str]] = None,
) -> List[Property]:
    """Get property information for an array of objects, flattened to `name#N.child` columns.

    `count` is the number of slots to generate (see `get_array_slot_count`).
    `required_keys` are child property names that should be marked required in
    every slot; the submission-schema endpoint does not propagate the nested
    object's own `required` list, so callers must supply it explicitly (see
    `get_object_required`).
    """
    object_properties = []
    required_keys = required_keys or []
    for index in range(0, count):
        for key, value in property_schema.items():
            combined_property_name = f"{property_name}#{index}.{key}"
            object_properties.append(
                get_property(
                    item,
                    combined_property_name,
                    value,
                    is_nested=True,
                    required_override=True if key in required_keys else None,
                )
            )
    return object_properties


def get_flattened_object_properties(
    item: str,
    property_name: str,
    object_properties: Dict[str, Any],
    required_keys: Optional[List[str]] = None,
) -> List[Property]:
    """Get property information for a plain (non-array) nested object, flattened to `name.child` columns.

    `required_keys` are child property names to mark required, sourced from the
    nested object's own `required` list (see `get_object_required`) since these
    children have no `is_required` annotation of their own.
    """
    required_keys = required_keys or []
    flattened_properties = []
    for key, value in object_properties.items():
        combined_property_name = f"{property_name}.{key}"
        flattened_properties.append(
            get_property(
                item,
                combined_property_name,
                value,
                required_override=True if key in required_keys else None,
            )
        )
    return flattened_properties


def get_object_required(object_schema: Dict[str, Any]) -> List[str]:
    """Get the `required` list declared directly on an object (or array `items`) schema.

    Used to propagate required-ness into flattened nested-object/array-of-object
    columns, since neither the submission-schema endpoint nor a raw profile
    schema's per-property `is_required`/absence thereof reflects this on its own.
    """
    return object_schema.get("required") or []


def get_plain_object_properties(property_schema: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Get the `properties` dict if the schema is a plain (non-array, non-link) object.

    Returns None for arrays (handled separately by `get_array_object_properties`),
    links, and scalars.
    """
    if property_schema.get("type") == "object" and not is_link(property_schema):
        return property_schema.get("properties") or None
    return None


def get_array_object_properties(property_schema: Dict[str, Any]) -> Union[Dict[str,Any], None]:
    """Get nested properties if property is an array of objects."""
    if item := property_schema.get("items",""):
        return item.get("properties","")
    return ""


def is_required(property_schema: Dict[str, Any]) -> bool:
    """Check if property is required"""
    return property_schema.get(SubmissionSchemaConstants.IS_REQUIRED, False)


def get_search_url(property_schema: Dict[str, Any]) -> str:
    """Get portal search url for linked item names."""
    if is_link(property_schema):
        linked_item = get_linkto(property_schema) or get_linkto(schema_utils.get_items(property_schema))
        return f"https://data.smaht.org/search/?type={linked_item}"
    else:
        return ""


def get_linkto(property_schema: Dict[str, Any]) -> str:
    """Get item type a property links to."""
    return property_schema.get("linkTo","")


def is_link(property_schema: Dict[str, Any]) -> bool:
    """Check if property is a link to another item"""
    return schema_utils.is_link(property_schema) or is_array_of_links(property_schema)


def is_array_of_links(property_schema: Dict[str, Any]) -> bool:
    """Check if property is an array of links"""
    return schema_utils.is_link(schema_utils.get_items(property_schema))


def get_enum(property_schema: Dict[str, Any]) -> List[str]:
    """Get the enum values"""
    return schema_utils.get_enum(property_schema) or get_nested_enum(property_schema)


def get_nested_enum(property_schema: Dict[str, Any]) -> List[str]:
    """Get the enum values from a nested schema"""
    return schema_utils.get_enum(schema_utils.get_items(property_schema))


def get_array_subtype(property_schema: Dict[str, Any]) -> str:
    """Get the array subtype"""
    return schema_utils.get_schema_type(schema_utils.get_items(property_schema))


def get_suggested_enum(property_schema: Dict[str, Any]) -> List[str]:
    """Get suggested_enum or nested suggested_enum for property values."""

    return schema_utils.get_suggested_enum(
        property_schema
    ) or schema_utils.get_suggested_enum(schema_utils.get_items(property_schema))


def get_examples(property_schema: Dict[str, Any]) -> List[str]:
    """Get examples for property values."""
    return schema_utils.get_submission_examples(
        property_schema
    ) or get_suggested_enum(property_schema)


def get_corequirements(property_schema: Dict[str, Any]) -> List[str]:
    """Get the corequirements for the property."""
    return property_schema.get(SubmissionSchemaConstants.ALSO_REQUIRES) or []


def get_exclusive_requirements(property_schema: Dict[str, Any]) -> List[str]:
    """Get the exclusive requirements for the property."""
    return property_schema.get(SubmissionSchemaConstants.REQUIRED_IF_NOT_ONE_OF) or []


def is_allow_commas(property_schema: Dict[str, Any]) -> bool:
    """Check if allow_commas is present in the property."""
    return property_schema.get("allow_commas", False)


def is_allow_multiplier_suffix(property_schema: Dict[str, Any]) -> bool:
    """Check if allow_multiplier is present in the property."""
    return property_schema.get("allow_multiplier_suffix", False)


def write_spreadsheet(
    output: Path, spreadsheet: Spreadsheet, separate_comments: bool = False, example: bool = False
) -> None:
    """Write spreadsheet to file"""
    file_path = get_output_file_path(output, spreadsheet)
    workbook = generate_workbook(spreadsheet, separate_comments=separate_comments)
    save_workbook(workbook, file_path)
    if example:
        log.info(f"Example spreadsheet written to: {file_path}")
    else:
        log.info(f"Spreadsheet written to: {file_path}")
    

def get_output_file_path(output: Path, spreadsheet: Spreadsheet) -> Path:
    """Get the output file path"""
    return Path(output, f"{to_snake_case(spreadsheet.item)}{ITEM_SPREADSHEET_SUFFIX}")


def generate_workbook(
    spreadsheet: Spreadsheet, separate_comments: bool = False
) -> openpyxl.Workbook:
    """Generate the workbook"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    set_sheet_name(worksheet, spreadsheet)
    write_properties(
        worksheet, spreadsheet.properties, separate_comments=separate_comments,examples=spreadsheet.examples
    )
    return workbook


def set_sheet_name(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, spreadsheet: Spreadsheet
) -> None:
    """Set the sheet name"""
    worksheet.title = to_camel_case(spreadsheet.item)


def write_properties(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    properties: List[Property],
    separate_comments: bool = False,
    examples: Optional[List[Dict[str,Any]]] = None
) -> None:
    """Write properties to the worksheet"""
    ordered_properties = get_ordered_properties(properties)
    for index, property_ in enumerate(ordered_properties, start=1):  # cells 1-indexed
        if separate_comments:
            write_property(worksheet, index, property_, comments=False)
            write_comment_cells(worksheet, index, property_)
        else:
            write_property(worksheet, index, property_)
        if examples:
            for row, item in enumerate(examples, start=2):
                if property_.nested:
                    value = get_nested_example(item,property_.name,)
                    write_example(worksheet,index,row,property_,value)
                elif property_.name in item:
                    value = item[property_.name]
                    write_example(worksheet,index,row,property_,value)
                else:
                    write_empty(worksheet,index,row,property_)


def extract_nested_property_names(property_name: str):
    """Extract the parent and nested property names from a nested property column name.
    
    Expects the format parent_property#n_index.nested_property """
    parent_property = property_name.split("#")[0]
    nested_property = property_name.split(".")[1]
    n_index= int(property_name.split("#")[1].split(".")[0])
    return parent_property, nested_property, n_index


def get_nested_example(item: Dict[str, Any],property_name: str):
    """Get example values from properties nested within an array of objects."""
    parent_property, nested_property, n_index = extract_nested_property_names(property_name)
    return item[parent_property][n_index][nested_property]


def get_ordered_properties(properties: List[Property]) -> List[Property]:
    """Order properties to write.

    Rank via:
       - Required non-links (~alphabetically)
       - Non-required non-links (not alphabetically)
       - Required links (alphabetically)
       - Non-required links (alphabetically)

    For arrays of objects, these may still be out of order
    """
    return [
        *get_required_non_links(properties),
        *get_non_required_non_links(properties),
        *get_required_links(properties),
        *get_non_required_links(properties),
    ]


def sort_properties_alphabetically(properties: List[Property]) -> List[Property]:
    """Sort properties alphabetically."""
    return sorted(properties, key=lambda property_: property_.name)


def get_required_non_links(properties: List[Property]) -> List[Property]:
    """Get required non-link properties.

    Separate out `submitted_id` if present as first property, and sort
    the rest alphabetically.
    """
    required_non_links = [
        property_
        for property_ in properties
        if property_.required and not property_.link
    ]
    submitted_id = [
        property_ for property_ in required_non_links if is_submitted_id(property_)
    ]
    non_submitted_id = sort_properties_alphabetically(
        [
            property_
            for property_ in required_non_links
            if not is_submitted_id(property_)
        ]
    )
    return [*submitted_id, *non_submitted_id]


def get_non_required_non_links(properties: List[Property]) -> List[Property]:
    """Get non-required non-link properties.
    
    No longer sort alphabetically.
    """
    return [
        property_
        for property_ in properties
        if not property_.required and not property_.link
    ]


def get_required_links(properties: List[Property]) -> List[Property]:
    """Get required link properties."""
    return [property_ for property_ in properties if property_.required and property_.link]


def get_non_required_links(properties: List[Property]) -> List[Property]:
    """Get non-required link properties."""
    return sort_properties_alphabetically(
        [
            property_
            for property_ in properties
            if not property_.required and property_.link
        ]
    )


def is_submitted_id(property_: Property) -> bool:
    """Check if property is `submitted_id`."""
    return property_.name == item_constants.SUBMITTED_ID


def write_example(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    index: int,
    row: int,
    property_: Property,
    value: Any,
) -> None:
    """Write example of property value to the worksheet"""
    if is_list(value):
        value=get_example_list(value)
    row =  row  # cells 1-indexed, start with 2
    cell = worksheet.cell(row=row, column=index, value=value)
    cell.font = openpyxl.styles.Font(name=FONT, size=FONT_SIZE)


def write_empty(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    index: int,
    row: int,
    property_: Property
):
    """Write empty cell to the worksheet."""
    cell = worksheet.cell(row=row, column=index, value=None)
    cell.font = openpyxl.styles.Font(name=FONT, size=FONT_SIZE)


def is_list(value):
    """Returns True if value is a list."""
    return type(value) is list


def get_example_list(value: List[Any]) -> List[str]:
    """"Convert list of values into | separate string."""
    return ' | '.join(value)

def write_property(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    index: int,
    property_: Property,
    comments: bool = True,
) -> None:
    """Write property to the worksheet"""
    row = 1  # cells 1-indexed
    cell = worksheet.cell(row=row, column=index, value=property_.name)
    set_cell_font(cell, property_)
    set_cell_width(worksheet, index, property_)
    if comments:
        write_comment(worksheet, index, property_)


def set_cell_font(cell: openpyxl.cell.cell.Cell, property_: Property) -> None:
    """Set the font for the cell."""
    font = get_font(property_)
    cell.font = font


def set_cell_width(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, index: int, property_: Property
) -> None:
    """Set the width of the cell."""
    min_width = 13  # Default width from openpyxl; looks reasonable
    calculated_width = len(property_.name) + 2
    width = calculated_width if calculated_width > min_width else min_width
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width


def write_comment(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, index: int, property_: Property
) -> None:
    """Write comment to the worksheet."""
    row = 1
    comment = get_comment(property_)
    if comment:
        worksheet.cell(row=row, column=index).comment = comment


def write_comment_cells(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, index: int, property_: Property
) -> None:
    """Write comment to separate cells under properties."""
    row = 2
    comment = get_comment_text(property_)
    if comment:
        worksheet.cell(row=row, column=index, value=comment)


def get_font(property_: Property) -> openpyxl.styles.Font:
    """Get font for the property."""
    font = openpyxl.styles.Font(name=FONT, size=FONT_SIZE)
    if property_.required:
        font.bold = True
    if property_.link:
        font.italic = True
    return font


def get_comment(property_: Property) -> Union[openpyxl.comments.Comment, None]:
    """Get comment for the property."""
    comment_text = get_comment_text(property_)
    if comment_text:
        height = get_comment_height(comment_text)
        width = get_comment_width(comment_text)
        return openpyxl.comments.Comment(comment_text, "", height=height, width=width)
    return None


def get_comment_text(property_: Property) -> str:
    """Get comment text for the property.

    Order of lines as defined here will be the order in the comment.
    """
    comment_lines = []
    indent = "  "
    comment_lines += get_comment_description(property_, indent)
    comment_lines += get_comment_value_type(property_, indent)
    comment_lines += get_comment_numbers(property_, indent)
    comment_lines += get_comment_enum(property_, indent)
    comment_lines += get_comment_examples(property_, indent)
    comment_lines += get_comment_link(property_, indent)
    comment_lines += get_comment_required(property_, indent)
    comment_lines += get_comment_requires(property_, indent)
    comment_lines += get_comment_pattern(property_, indent)
    comment_lines += get_comment_note(property_, indent)
    comment_lines += get_comment_search(property_, indent)
    comment_lines += get_comment_nested(property_, indent)

    return "\n".join(comment_lines)


def get_comment_description(property_: Property, indent: str) -> List[str]:
    if property_.description:
        return [f"Description:{indent}{property_.description}"]
    return []


def get_comment_value_type(property_: Property, indent: str) -> List[str]:
    if property_.value_type:
        if property_.array_subtype:
            return [
                (
                    f"Type:{indent}{property_.array_subtype}"
                    f"{indent}(Multiple values allowed. Use '|' as a delimiter.)"
                )
            ]
        else:
            return [f"Type:{indent}{property_.value_type}"]
    return []


def get_comment_enum(property_: Property, indent: str) -> List[str]:
    if property_.enum:
        return [f"Options:{indent}{' | '.join(property_.enum)}"]
    return []


def get_comment_examples(property_: Property, indent: str) -> List[str]:
    if property_.examples:
        return [f"Examples:{indent}{' | '.join(property_.examples)}"]
    return []


def get_comment_link(property_: Property, indent: str) -> List[str]:
    if property_.link:
        return [f"Link:{indent}Yes"]
    return []


def get_comment_nested(property_: Property, indent: str) -> List[str]:
    if property_.nested:
        return [f"Nested:{indent}Yes"]
    return []


def get_comment_search(property_: Property, indent: str) -> List[str]:
    """Get text for search url.
    
    If property is file_format, include query for specific File type
    """
    search_text = "Use URL to search for the submitted_id or identifer of relevant items"
    if property_.search:
        if property_.name == "file_format":
            return [f"{search_text}:{indent}{property_.search}&valid_item_types={property_.item}"]
        return [f"{search_text}:{indent}{property_.search}"]
    return []


def get_comment_required(property_: Property, indent: str) -> List[str]:
    if property_.required:
        return [f"Required:{indent}Yes"]
    if property_.exclusive_requirements:
        return [
            (
                f"Required:{indent}Possibly\n"
                f"{indent}Not required if present:{indent}"
                f"{' | '.join(property_.exclusive_requirements)}"
            )
        ]
    return [f"Required:{indent}No"]


def get_comment_requires(property_: Property, indent: str) -> List[str]:
    if property_.requires:
        return [f"Requires:{indent}{' | '.join(property_.requires)}"]
    return []


def get_comment_pattern(property_: Property, indent: str) -> List[str]:
    if property_.pattern:
        return [f"Pattern:{indent}{property_.pattern}"]
    if is_date_format(property_.format_):
        return [f"Format:{indent}YYYY-MM-DD"]
    return []


def get_comment_note(property_: Property, indent: str) -> List[str]:
    if property_.comment:
        return [f"Note:{indent}{property_.comment}"]
    return []


def get_comment_numbers(property_: Property, indent: str) -> List[str]:
    """Get comment for allow_commas and allow_multiplier_suffix."""
    comment = []
    if property_.allow_commas:
        comment.append(f"{indent} Commas allowed (e.g. 1,000,000)")
    if property_.allow_multiplier_suffix:
        comment.append(f"{indent} Abbreviations allowed, such as 1M for 1000000 or 10.5kb for 10500 bp")
    return comment


def is_date_format(format_: str) -> bool:
    """Check if the format is a date format."""
    return format_ == "date"


def get_comment_height(comment_text: str) -> int:
    """Get comment height based on number of lines."""
    pixels_per_line = 25  # Looks reasonable for the font size
    lines = [line for line in comment_text.split("\n")]
    return len(lines) * pixels_per_line


def get_comment_width(comment_text: str) -> int:
    """Get comment width based on longest line.

    Note: Pixels per character chosen based on trial and error.
    """
    lines = [line for line in comment_text.split("\n")]
    pixels_per_char = 6
    return max(len(line) for line in lines) * pixels_per_char


def save_workbook(workbook: openpyxl.Workbook, file_path: Path) -> None:
    """Save the workbook to the file path"""
    workbook.save(filename=file_path)


# --- Publication / StaticSection workbook ---
#
# Attaches StaticSection content blocks to existing Publications. Unlike the rest
# of this module, this reads the raw `/profiles/<item>.json` schema rather than
# `/submission-schemas/<item>.json`: the submission-schema endpoint unconditionally
# drops `accession`, `uuid`, `submission_centers`, and `consortia` (see
# `SMaHTProjectSchemaViews.get_properties_for_exclusion`) and also drops the
# `required` list of nested array-item objects (see
# `snovault.schema_views._build_embedded_obj`), all of which this admin-facing
# workbook needs. Since the raw schema carries no `is_required` annotation at all,
# required-ness here is always derived explicitly from the schema's own `required`
# lists via `required_override`, never from `is_required()`.

PUBLICATION_STATIC_SECTION_WORKBOOK_FILENAME = "publication_static_section_submission.xlsx"
DEFAULT_STATIC_CONTENT_SLOTS = 3
DEFAULT_STATIC_SECTION_LOCATIONS = ["key-findings", "reference-set-generation", "header"]
DEFAULT_SECTION_TYPE_ENUM = [
    "Page Section", "Announcement", "Search Info Header", "Item Page Header", "Home Page Slide"
]
DEFAULT_FILETYPE_ENUM = ["md", "html", "txt", "csv", "jsx", "rst"]
BOOLEAN_DROPDOWN_VALUES = ["true", "false"]
DEFAULT_DROPDOWN_DATA_ROWS = 200  # dropdown validation is applied to this many blank rows in template mode

OVERVIEW_SHEET_NAME = "(Overview Guidelines)"
DROPDOWN_HELPER_SHEET_NAME = "(Dropdown Options)"
LOCATIONS_HELPER_SHEET_NAME = "(Locations)"

STATIC_SECTION_ITEM = "StaticSection"
PUBLICATION_ITEM = "Publication"

STATIC_SECTION_CURATED_PROPERTIES = [
    "identifier", "title", "body", "file", "section_type",
    "options.filetype", "options.collapsible", "options.default_open",
    "options.title_icon", "options.link", "options.image",
    "description", "submission_centers", "consortia",
]


def get_raw_profile_schema(item: str, request_handler: RequestHandler) -> Dict[str, Any]:
    """Get the full (non-submission-filtered) profile schema for an item.

    See module note above for why this, rather than `get_submission_schema`, is
    used for the Publication/StaticSection workbook.
    """
    return request_handler.get_item(f"/profiles/{to_snake_case(item)}.json")


def get_publication_curated_property_names(
    static_content_slots: int, include_static_headers: bool
) -> List[str]:
    """Get the curated Publication sheet column names for the configured slot count."""
    names = ["accession", "uuid"]
    for index in range(static_content_slots):
        names += [
            f"static_content#{index}.content",
            f"static_content#{index}.location",
            f"static_content#{index}.description",
        ]
    if include_static_headers:
        names.append("static_headers")
    return names


def get_curated_properties(
    item: str,
    schema: Dict[str, Any],
    curated_names: List[str],
    array_slot_counts: Optional[Dict[str, int]] = None,
) -> List[Property]:
    """Build a curated Property list from a raw profile schema.

    Generates the full Property list for the schema (flattening nested objects
    and arrays of objects per `get_nested_properties`), then keeps only the
    properties named in `curated_names` — this workbook intentionally exposes a
    fixed, hand-picked subset of each schema's fields, not every submittable
    property.
    """
    top_required = get_object_required(schema)
    properties = []
    for key, value in schema_utils.get_properties(schema).items():
        properties += get_nested_properties(
            item, key, value,
            array_slot_counts=array_slot_counts,
            required_override=key in top_required,
        )
    return [property_ for property_ in properties if property_.name in curated_names]


def get_property_by_name(properties: List[Property], name: str) -> Optional[Property]:
    """Get the first property matching `name`, if any."""
    for property_ in properties:
        if property_.name == name:
            return property_
    return None


def append_comment_note(existing_comment: str, note: str) -> str:
    """Append a note to an existing comment string, if any."""
    return f"{existing_comment} {note}".strip() if existing_comment else note


def get_replace_semantics_note(append: bool) -> str:
    """Get the note explaining --append/--replace update semantics."""
    if append:
        return (
            "Append mode: in populated mode, each Publication's existing static_content"
            " entries are pre-filled into the first slots; fill in the remaining empty"
            " slots to add new StaticSections without losing the existing ones. On"
            " import, the full slot list (existing + new) replaces the item's"
            " static_content, since array properties are always submitted as a whole."
        )
    return (
        "Replace mode: existing static_content entries are NOT pre-filled, even in"
        " populated mode. The complete list you enter becomes the item's entire"
        " static_content on import; any existing entries you don't re-list are removed."
    )


def augment_static_section_properties(properties: List[Property]) -> List[Property]:
    """Add workbook-specific comment notes to StaticSection properties that the schema alone can't express."""
    augmented = []
    for property_ in properties:
        if property_.name == "identifier":
            property_ = dataclasses.replace(
                property_,
                comment=append_comment_note(
                    property_.comment,
                    "Convention used in this workbook: <Publication accession>.<location>,"
                    " e.g. SMAHT001.key-findings.",
                ),
            )
        elif property_.name in ("body", "file"):
            property_ = dataclasses.replace(
                property_,
                comment=append_comment_note(
                    property_.comment, "Provide at most one of `body`/`file`, not both."
                ),
            )
        elif property_.name in ("submission_centers", "consortia"):
            property_ = dataclasses.replace(
                property_,
                comment=append_comment_note(
                    property_.comment,
                    "As an admin submitter, leave blank to have this added automatically;"
                    " set explicitly only to override the default attribution.",
                ),
            )
        augmented.append(property_)
    return augmented


def augment_publication_properties(
    properties: List[Property], locations: List[str], append: bool
) -> List[Property]:
    """Add workbook-specific comment notes/enum overrides to Publication properties."""
    replace_semantics = get_replace_semantics_note(append)
    augmented = []
    for property_ in properties:
        if property_.name == "accession":
            property_ = dataclasses.replace(
                property_,
                # Not required by the Publication schema itself (accession is
                # server-assigned); required here because it's how this workbook
                # locates the existing Publication to update.
                required=True,
                comment=append_comment_note(
                    property_.comment,
                    "Preferred identifier for locating the existing Publication to update.",
                ),
            )
        elif property_.name == "uuid":
            property_ = dataclasses.replace(
                property_,
                comment=append_comment_note(
                    property_.comment, "Alternative identifier if accession is not known."
                ),
            )
        elif property_.name.endswith(".content"):
            property_ = dataclasses.replace(
                property_,
                comment=append_comment_note(
                    append_comment_note(
                        property_.comment,
                        "Provide the `identifier` or `uuid` of the target StaticSection item.",
                    ),
                    replace_semantics,
                ),
            )
        elif property_.name.endswith(".location"):
            property_ = dataclasses.replace(
                property_,
                enum=list(locations),
                comment=append_comment_note(
                    append_comment_note(
                        property_.comment,
                        f"Choose from the configured locations (see '{LOCATIONS_HELPER_SHEET_NAME}' sheet).",
                    ),
                    replace_semantics,
                ),
            )
        elif property_.name == "static_headers":
            property_ = dataclasses.replace(
                property_,
                comment=append_comment_note(
                    append_comment_note(
                        property_.comment,
                        "Pipe-delimited `identifier`/`uuid` values of StaticSection items to"
                        " show at the top of the Publication page.",
                    ),
                    replace_semantics,
                ),
            )
        augmented.append(property_)
    return augmented


def write_curated_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet, properties: List[Property]
) -> List[Property]:
    """Write a curated property list to a worksheet's header row and return the written order.

    Reuses `write_property`/`get_ordered_properties` so column ordering (required
    non-links, optional non-links, required links, optional links), fonts, widths,
    and comments exactly match the conventions used everywhere else in this module.
    """
    ordered_properties = get_ordered_properties(properties)
    for index, property_ in enumerate(ordered_properties, start=1):
        write_property(worksheet, index, property_)
    return ordered_properties


def write_overview_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    static_content_slots: int,
    locations: List[str],
    mode: str,
    append: bool,
    include_static_headers: bool,
) -> None:
    """Write plain-text usage guidelines to the overview sheet.

    This sheet's name is parenthesized so submitr ignores it on import; unlike the
    true helper sheets it stays visible, since it's meant to be read by whoever
    fills in the workbook.
    """
    mode_description = (
        "headers/comments/dropdowns only, no data rows"
        if mode == "template"
        else "pre-filled with existing Publications from the portal"
    )
    lines = [
        "Guidelines for attaching StaticSection content to existing Publications.",
        "",
        f"Mode: {mode} ({mode_description}).",
        "",
        "Sheet order:",
        f"  1. {OVERVIEW_SHEET_NAME} - this sheet; ignored by submitr (parenthesized name).",
        f"  2. {STATIC_SECTION_ITEM} - create/describe the StaticSection content block(s).",
        f"  3. {PUBLICATION_ITEM} - attach StaticSection(s) to an existing Publication"
        " (identified by accession or uuid).",
        "  4. Hidden helper sheets - dropdown source data; ignored by submitr"
        " (hidden + parenthesized names).",
        "",
        "StaticSection identifier convention used in this workbook:",
        "  <Publication accession>.<location>, e.g. SMAHT001.key-findings",
        "",
        f"Configured static_content slots: {static_content_slots}",
        f"Configured locations: {' | '.join(locations)}",
        f"static_headers column included: {include_static_headers}",
        "",
        "Update semantics (--append / --replace):",
        f"  {get_replace_semantics_note(append)}",
        "",
        "submission_centers / consortia on StaticSection:",
        "  As an admin submitter, leave these blank to have them added automatically;"
        " set explicitly only to override the default attribution.",
    ]
    for row, line in enumerate(lines, start=1):
        cell = worksheet.cell(row=row, column=1, value=line)
        cell.font = openpyxl.styles.Font(name=FONT, size=FONT_SIZE)
    worksheet.column_dimensions["A"].width = 110


def write_dropdown_helper_sheet(
    workbook: openpyxl.Workbook, section_type_enum: List[str], filetype_enum: List[str]
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Write the hidden helper sheet backing the fixed-enum/boolean dropdowns."""
    worksheet = workbook.create_sheet(title=DROPDOWN_HELPER_SHEET_NAME)
    columns = {
        "A": ("Section Type", section_type_enum),
        "B": ("File Type", filetype_enum),
        "C": ("Boolean", BOOLEAN_DROPDOWN_VALUES),
    }
    for column, (header, values) in columns.items():
        worksheet[f"{column}1"] = header
        for offset, value in enumerate(values, start=2):
            worksheet[f"{column}{offset}"] = value
    worksheet.sheet_state = "hidden"
    return worksheet


def write_locations_helper_sheet(
    workbook: openpyxl.Workbook, locations: List[str]
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Write the hidden helper sheet backing the user-configured location dropdown."""
    worksheet = workbook.create_sheet(title=LOCATIONS_HELPER_SHEET_NAME)
    worksheet["A1"] = "Location"
    for offset, location in enumerate(locations, start=2):
        worksheet[f"A{offset}"] = location
    worksheet.sheet_state = "hidden"
    return worksheet


def get_dropdown_range(sheet_name: str, column: str, count: int) -> str:
    """Get a quoted absolute range reference into a helper sheet's dropdown column."""
    return f"'{sheet_name}'!${column}$2:${column}${max(count, 1) + 1}"


def add_dropdown_validation(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    column_index: int,
    range_reference: str,
    data_row_count: int,
) -> None:
    """Attach a list-type data validation to a column's data rows (below the header)."""
    if data_row_count < 1:
        return
    validation = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1=f"={range_reference}", allow_blank=True
    )
    worksheet.add_data_validation(validation)
    column_letter = openpyxl.utils.get_column_letter(column_index)
    validation.add(f"{column_letter}2:{column_letter}{data_row_count + 1}")


def apply_static_section_dropdowns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    ordered_properties: List[Property],
    section_type_enum: List[str],
    filetype_enum: List[str],
    data_row_count: int,
) -> None:
    """Attach dropdowns for StaticSection's fixed-enum/boolean columns."""
    section_type_range = get_dropdown_range(DROPDOWN_HELPER_SHEET_NAME, "A", len(section_type_enum))
    filetype_range = get_dropdown_range(DROPDOWN_HELPER_SHEET_NAME, "B", len(filetype_enum))
    boolean_range = get_dropdown_range(DROPDOWN_HELPER_SHEET_NAME, "C", len(BOOLEAN_DROPDOWN_VALUES))
    for index, property_ in enumerate(ordered_properties, start=1):
        if property_.name == "section_type":
            add_dropdown_validation(worksheet, index, section_type_range, data_row_count)
        elif property_.name == "options.filetype":
            add_dropdown_validation(worksheet, index, filetype_range, data_row_count)
        elif property_.name in ("options.collapsible", "options.default_open"):
            add_dropdown_validation(worksheet, index, boolean_range, data_row_count)


def apply_publication_dropdowns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    ordered_properties: List[Property],
    locations: List[str],
    data_row_count: int,
) -> None:
    """Attach dropdowns for Publication's user-configured location columns."""
    locations_range = get_dropdown_range(LOCATIONS_HELPER_SHEET_NAME, "A", len(locations))
    for index, property_ in enumerate(ordered_properties, start=1):
        if property_.name.endswith(".location"):
            add_dropdown_validation(worksheet, index, locations_range, data_row_count)


def get_existing_publications(
    request_handler: RequestHandler, portal_url: Optional[str] = None
) -> List[Dict[str, Any]]:
    """Search the portal for existing Publications, for populated mode."""
    search = "search/?type=Publication&frame=object"
    auth_key = dict(request_handler.auth_key) if request_handler.auth_key else {}
    if portal_url:
        auth_key = {**auth_key, "server": portal_url}
    try:
        return ff_utils.search_metadata(search, key=auth_key)
    except Exception as e:
        log.error(f"Error fetching existing Publications: {e}")
        return []


def get_publication_slot_count(ordered_properties: List[Property]) -> int:
    """Get the configured number of static_content slots from an ordered Publication property list."""
    return len({
        extract_nested_property_names(property_.name)[2]
        for property_ in ordered_properties
        if property_.nested and property_.name.startswith("static_content#")
    })


def write_existing_publication_rows(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    ordered_properties: List[Property],
    publications: List[Dict[str, Any]],
    append: bool,
    request_handler: RequestHandler,
) -> None:
    """Pre-fill one row per existing Publication: identifier columns always, static_content only if appending."""
    column_by_name = {property_.name: index for index, property_ in enumerate(ordered_properties, start=1)}
    accession_column = column_by_name.get("accession")
    uuid_column = column_by_name.get("uuid")
    slot_count = get_publication_slot_count(ordered_properties)
    for row, publication in enumerate(publications, start=2):
        if accession_column:
            worksheet.cell(row=row, column=accession_column, value=publication.get("accession", ""))
        if uuid_column:
            worksheet.cell(row=row, column=uuid_column, value=publication.get("uuid", ""))
        if append:
            write_existing_static_content(worksheet, column_by_name, publication, row, slot_count, request_handler)


def write_existing_static_content(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    column_by_name: Dict[str, int],
    publication: Dict[str, Any],
    row: int,
    slot_count: int,
    request_handler: RequestHandler,
) -> None:
    """Pre-fill a Publication's existing static_content entries into the configured slots.

    If the Publication has more existing entries than configured slots, the extras
    are logged and skipped rather than silently dropped.
    """
    existing = publication.get("static_content") or []
    if len(existing) > slot_count:
        log.warning(
            f"Publication {publication.get('accession') or publication.get('uuid')} has"
            f" {len(existing)} existing static_content entries but only {slot_count} slots"
            f" are configured; the last {len(existing) - slot_count} will not be pre-filled."
            f" Increase --static-content-slots to include them."
        )
    for index, entry in enumerate(existing[:slot_count]):
        content_column = column_by_name.get(f"static_content#{index}.content")
        location_column = column_by_name.get(f"static_content#{index}.location")
        description_column = column_by_name.get(f"static_content#{index}.description")
        if content_column:
            worksheet.cell(
                row=row, column=content_column,
                value=get_existing_static_content_identifier(entry.get("content", ""), request_handler),
            )
        if location_column:
            worksheet.cell(row=row, column=location_column, value=entry.get("location", ""))
        if description_column and entry.get("description"):
            worksheet.cell(row=row, column=description_column, value=entry.get("description"))


def get_existing_static_content_identifier(content: str, request_handler: RequestHandler) -> str:
    """Resolve a static_content entry's linked StaticSection to its `identifier`/`uuid`.

    `content` as returned by the portal (frame=object) is typically an `@id` path
    rather than the `identifier`/`uuid` this workbook expects submitters to enter,
    so it's resolved here to keep populated-mode output directly re-submittable.
    Falls back to the raw value if resolution fails.
    """
    if not content:
        return content
    try:
        static_section = request_handler.get_item(content)
    except Exception as e:
        log.error(f"Error resolving StaticSection {content}: {e}")
        return content
    return get_linked_item_id(static_section) or content


def get_property_enum(properties: List[Property], name: str, default: List[str]) -> List[str]:
    """Get a named property's enum values, falling back to `default` if absent/empty."""
    property_ = get_property_by_name(properties, name)
    return list(property_.enum) if property_ and property_.enum else list(default)


def write_publication_static_section_workbook(
    output: Path,
    request_handler: RequestHandler,
    static_content_slots: int = DEFAULT_STATIC_CONTENT_SLOTS,
    locations: Optional[List[str]] = None,
    mode: str = "template",
    append: bool = True,
    portal_url: Optional[str] = None,
    include_static_headers: bool = True,
) -> None:
    """Write a workbook for attaching StaticSection content to existing Publications.

    Sheet order: `(Overview Guidelines)` (visible), `StaticSection` (visible),
    `Publication` (visible), then hidden helper sheets backing the dropdowns.

    `mode="template"` writes headers/comments/dropdowns only. `mode="populated"`
    additionally fetches existing Publications (optionally from `portal_url`) and
    writes one row per Publication, pre-filling `accession`/`uuid` always and
    existing `static_content` only if `append` is True (see
    `get_replace_semantics_note`).
    """
    if mode not in ("template", "populated"):
        raise ValueError(f"Invalid mode: {mode}. Must be 'template' or 'populated'.")
    locations = list(locations) if locations else list(DEFAULT_STATIC_SECTION_LOCATIONS)

    static_section_schema = get_raw_profile_schema(STATIC_SECTION_ITEM, request_handler)
    publication_schema = get_raw_profile_schema(PUBLICATION_ITEM, request_handler)

    static_section_properties = get_curated_properties(
        STATIC_SECTION_ITEM, static_section_schema, STATIC_SECTION_CURATED_PROPERTIES
    )
    static_section_properties = augment_static_section_properties(static_section_properties)

    publication_curated_names = get_publication_curated_property_names(
        static_content_slots, include_static_headers
    )
    publication_properties = get_curated_properties(
        PUBLICATION_ITEM, publication_schema, publication_curated_names,
        array_slot_counts={"static_content": static_content_slots},
    )
    publication_properties = augment_publication_properties(publication_properties, locations, append)

    section_type_enum = get_property_enum(static_section_properties, "section_type", DEFAULT_SECTION_TYPE_ENUM)
    filetype_enum = get_property_enum(static_section_properties, "options.filetype", DEFAULT_FILETYPE_ENUM)

    workbook = openpyxl.Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = OVERVIEW_SHEET_NAME
    write_overview_sheet(
        overview_sheet, static_content_slots, locations, mode, append, include_static_headers
    )

    static_section_sheet = workbook.create_sheet(title=STATIC_SECTION_ITEM)
    ordered_static_section_properties = write_curated_sheet(static_section_sheet, static_section_properties)

    publication_sheet = workbook.create_sheet(title=PUBLICATION_ITEM)
    ordered_publication_properties = write_curated_sheet(publication_sheet, publication_properties)

    write_dropdown_helper_sheet(workbook, section_type_enum, filetype_enum)
    write_locations_helper_sheet(workbook, locations)

    data_row_count = DEFAULT_DROPDOWN_DATA_ROWS
    existing_publications = []
    if mode == "populated":
        existing_publications = get_existing_publications(request_handler, portal_url)
        # Extend validation past the pre-filled rows too, so manually-added rows
        # (e.g. for a Publication not returned by the search) still get dropdowns.
        data_row_count = len(existing_publications) + DEFAULT_DROPDOWN_DATA_ROWS

    apply_static_section_dropdowns(
        static_section_sheet, ordered_static_section_properties, section_type_enum, filetype_enum, data_row_count
    )
    apply_publication_dropdowns(publication_sheet, ordered_publication_properties, locations, data_row_count)

    if mode == "populated" and existing_publications:
        write_existing_publication_rows(
            publication_sheet, ordered_publication_properties, existing_publications, append, request_handler
        )

    save_workbook(workbook, output)
    log.info(f"Publication/StaticSection workbook ({mode} mode) written to: {output}")


def main():
    parser = argparse.ArgumentParser(description="Write submission spreadsheets")
    parser.add_argument(
        "--output",
        help=(
            f"Output location. For every mode except --publication-static-sections,"
            f" this is an existing output file *directory*: if not creating a"
            f" workbook, separate files will be created for each item with the"
            f" filename <item>{ITEM_SPREADSHEET_SUFFIX}; if creating a workbook,"
            f" the workbook will be saved as {WORKBOOK_FILENAME}."
            f" For --publication-static-sections, this is instead the exact output"
            f" *file* path, defaulting to {PUBLICATION_STATIC_SECTION_WORKBOOK_FILENAME}."
        ),
    )
    parser.add_argument("--env", help="Environment", default="data")
    parser.add_argument("--item", help="Item name", nargs="+")
    parser.add_argument("--tpc", help="TPC Submission items", action="store_true")
    parser.add_argument("--gcc", help="GCC Submission items", action="store_true")
    parser.add_argument("--all", help="All items", action="store_true")
    parser.add_argument(
        "--workbook",
        help=(
            f"Bundle all items into a single workbook."
            f" Workbook will be saved as {WORKBOOK_FILENAME}."
        ),
        action="store_true",
    )
    parser.add_argument(
        "--separate", help="Add comments as separate cells", action="store_true"
    )
    parser.add_argument(
        "--google",
        help=(
            f"Google Sheets ID to write."
            f" Expects credentials in {GOOGLE_CREDENTIALS_PATH}."
            f" Token will be saved to {GOOGLE_TOKEN_PATH}."
            f" For more information, see docstring within the script."
        ),
    )
    parser.add_argument(
        "--eqm",
        choices=["dsa","duplexseq"],
        help="External Quality Metric for specific submissions"
    )
    parser.add_argument(
        "--example",
        help=(f" Write out populated example submission template."
        f"Currently works with --gcc and --item."
        f"Starts of with {EXAMPLE_FILE_UUIDS} as AlignedReads."
        ),
        action="store_true"
    )
    parser.add_argument(
        "--publication-static-sections",
        help=(
            "Write a workbook for attaching StaticSection content to existing"
            " Publications, instead of any of the normal per-item spreadsheet"
            " modes. Cannot be combined with --item/--gcc/--tpc/--all/--google/"
            "--eqm/--example."
        ),
        action="store_true",
    )
    parser.add_argument(
        "--static-content-slots",
        help=(
            "Number of `static_content#N.*` slot column groups to generate on the"
            f" Publication sheet. Only used with --publication-static-sections."
            f" (default: {DEFAULT_STATIC_CONTENT_SLOTS})"
        ),
        type=int,
        default=DEFAULT_STATIC_CONTENT_SLOTS,
    )
    parser.add_argument(
        "--locations",
        help=(
            "Location values to offer in the static_content location dropdown."
            " Only used with --publication-static-sections."
            f" (default: {DEFAULT_STATIC_SECTION_LOCATIONS})"
        ),
        nargs="+",
        default=None,
    )
    update_semantics_group = parser.add_mutually_exclusive_group()
    update_semantics_group.add_argument(
        "--append",
        help=(
            "In populated mode, pre-fill each Publication's existing static_content"
            " entries into the first slots so new entries can be added alongside"
            " them. Only used with --publication-static-sections. (default)"
        ),
        dest="append",
        action="store_true",
        default=True,
    )
    update_semantics_group.add_argument(
        "--replace",
        help=(
            "In populated mode, leave static_content blank so the sheet's contents"
            " become the item's entire static_content on import, discarding any"
            " existing entries not re-listed. Only used with"
            " --publication-static-sections."
        ),
        dest="append",
        action="store_false",
    )
    parser.add_argument(
        "--portal-url",
        help=(
            "Portal server URL to search for existing Publications, overriding the"
            " server implied by --env. Giving --portal-url is what switches"
            " --publication-static-sections into populated mode (pre-filled from"
            " the portal); omitting it writes a blank template using --env only"
            " for schema lookups. Only used with --publication-static-sections."
        ),
    )
    static_headers_group = parser.add_mutually_exclusive_group()
    static_headers_group.add_argument(
        "--include-static-headers",
        help=(
            "Include the `static_headers` column on the Publication sheet. Only"
            " used with --publication-static-sections. (default: included)"
        ),
        dest="include_static_headers",
        action="store_true",
        default=True,
    )
    static_headers_group.add_argument(
        "--no-static-headers",
        help=(
            "Omit the `static_headers` column from the Publication sheet. Only used"
            " with --publication-static-sections."
        ),
        dest="include_static_headers",
        action="store_false",
    )
    args = parser.parse_args()

    keys = SMaHTKeyManager().get_keydict_for_env(args.env)
    log.info(f"Found keys for {args.env}")
    request_handler = RequestHandler(auth_key=keys)
    if args.publication_static_sections:
        if args.item or args.gcc or args.tpc or args.all or args.google or args.eqm or args.example:
            parser.error(
                "--publication-static-sections cannot be combined with"
                " --item/--gcc/--tpc/--all/--google/--eqm/--example"
            )
        if args.static_content_slots < 0:
            parser.error("--static-content-slots must be >= 0")
        output = Path(args.output) if args.output else Path(PUBLICATION_STATIC_SECTION_WORKBOOK_FILENAME)
        mode = "populated" if args.portal_url else "template"
        log.info(f"Writing Publication/StaticSection workbook ({mode} mode) to: {output}")
        write_publication_static_section_workbook(
            output,
            request_handler,
            static_content_slots=args.static_content_slots,
            locations=args.locations,
            mode=mode,
            append=args.append,
            portal_url=args.portal_url,
            include_static_headers=args.include_static_headers,
        )
        return
    if not args.output and not args.google:
        parser.error("No output specified")
    if args.output:
        args.output = dir_path(args.output)
    if args.gcc and args.tpc:
        parser.error("Cannot specify both gcc and tpc")
    if args.all and args.tpc:
        parser.error("Cannot specify both all and tpc")
    if args.all and args.gcc:
        parser.error("Cannot specify both all and gcc")
    if args.all and args.item:
        parser.error("Cannot specify both all and item")
    if args.eqm and args.example:
         parser.error("Currently cannot specify both eqm and example")
    if args.eqm and args.tpc:
        parser.error("Cannot specify both eqm and tpc")
    if args.eqm and not args.gcc and not args.item:
        parser.error("Need to specify gcc or item for ExternalQualityMetric output.\n"
        "Example: write-submission-spreadsheet --env [env] --output [output_path] --eqm dsa --item SupplementaryFile")
    if args.eqm:
        log.info(f"Writing ExternalQualityMetric spreadsheet for type: {args.eqm}")
    if args.example:
        if args.google:
            log.info(f"Google Sheet ID: {args.google}")
            log.info(f"Google Token Path: {GOOGLE_TOKEN_PATH}")
            spreadsheet_client = get_google_sheet_client(args.google)
            if args.gcc:
                log.info("Writing GCC submission example to Google sheet")
                update_google_sheets(spreadsheet_client, request_handler,gcc=True,example=True)
            elif args.item:
                log.info(f"Writing submission example to Google sheet for item(s): {args.item}")
                update_google_sheets(spreadsheet_client, request_handler,items=args.item,example=True)
        elif args.item:
            log.info(f"Writing example submission spreadsheets for item(s): {args.item}")
            write_item_spreadsheets(
                args.output,
                args.item,
                request_handler,
                workbook=args.workbook,
                separate_comments=args.separate,
                example=True
            )
        elif args.gcc:
            log.info("Writing example submission spreadsheet for GCC submission")
            write_item_spreadsheets(
                args.output,
                GCC_SUBMISSION_ITEMS,
                request_handler,
                workbook=args.workbook,
                separate_comments=args.separate,
                example=True,
                gcc=True
            )
        else: 
            parser.error("--example argument currently only works for individual item lists or --gcc")
    else:
        if args.google:
            # Google spreadsheets
            log.info(f"Google Sheet ID: {args.google}")
            log.info(f"Google Token Path: {GOOGLE_TOKEN_PATH}")
            spreadsheet_client = get_google_sheet_client(args.google)
            if args.all:
                update_google_sheets(spreadsheet_client, request_handler)
            elif args.gcc:
                log.info("Writing GCC submission Google sheet")
                update_google_sheets(spreadsheet_client, request_handler,gcc=True, eqm=args.eqm)
            elif args.tpc:
                log.info("Writing TPC submission Google sheet")
                update_google_sheets(spreadsheet_client, request_handler,tpc=True)
            elif args.item:
                log.info(f"Writing submission Google sheet for item(s): {args.item}")
                update_google_sheets(spreadsheet_client, request_handler,items=args.item, eqm=args.eqm)
            else:
                parser.error("No items specified to write or update Google spreadsheets for")
        elif args.all:
            log.info("Writing all submission spreadsheets")
            write_all_spreadsheets(
                args.output,
                request_handler,
                workbook=args.workbook,
                separate_comments=args.separate,
            )
        elif args.tpc:
            log.info("Writing TPC submission spreadsheet")
            write_item_spreadsheets(
                args.output,
                TPC_SUBMISSION_ITEMS,
                request_handler,
                workbook=args.workbook,
                tpc = True,
                gcc = False,
                separate_comments=args.separate,
            )
        elif args.gcc:
            log.info("Writing GCC/TTD submission spreadsheet")
            write_item_spreadsheets(
                args.output,
                GCC_SUBMISSION_ITEMS,
                request_handler,
                workbook=args.workbook,
                tpc = False,
                gcc = True,
                separate_comments=args.separate,
                eqm=args.eqm
            )
        elif args.item:
            log.info(f"Writing submission spreadsheets for item(s): {args.item}")
            write_item_spreadsheets(
                args.output,
                args.item,
                request_handler,
                workbook=args.workbook,
                separate_comments=args.separate,
                eqm=args.eqm
            )
        else:
            parser.error("No items specified to write or update spreadsheets for")


def dir_path(path: str) -> Path:
    """Check if the path is a directory"""
    if Path(path).is_dir():
        return Path(path)
    raise NotADirectoryError(path)


def get_google_sheet_client(sheet_id: str) -> SheetsClient:
    """Get Google Sheet to write/update."""
    credentials = get_google_credentials()
    service = build("sheets", "v4", credentials=credentials)
    return SheetsClient(service.spreadsheets(), sheet_id)


def get_google_credentials() -> Dict[str, Any]:
    """Get Google creds from secrets file."""
    creds = get_google_creds_from_token()
    if not creds:
        creds = get_google_creds_from_flow()
        write_google_token(creds)
    return creds


def get_google_creds_from_token() -> Union[Credentials, None]:
    """Get Google creds from token file."""
    if GOOGLE_TOKEN_PATH.exists():
        credentials = Credentials.from_authorized_user_file(
            GOOGLE_TOKEN_PATH, GOOGLE_SHEET_SCOPES
        )
        if credentials.valid:
            return credentials
    return None


def get_google_creds_from_flow() -> Credentials:
    """Get Google creds from flow via credentials file."""
    flow = InstalledAppFlow.from_client_secrets_file(
        GOOGLE_CREDENTIALS_PATH, GOOGLE_SHEET_SCOPES
    )
    return flow.run_local_server(port=0)


def write_google_token(creds: Credentials) -> None:
    """Write Google token to file."""
    with open(GOOGLE_TOKEN_PATH, "w") as file:
        file.write(creds.to_json())


if __name__ == "__main__":
    main()
