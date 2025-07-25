from typing import Dict, List, Any, Optional
import argparse
import openpyxl
import logging
from pathlib import Path

from encoded.commands.utils import get_auth_key
from encoded.item_utils.utils import (
    RequestHandler,
    get_property_values_from_identifiers
)

from encoded.item_utils import (
    item as item_utils,
    protected_donor as pd_utils,
    donor as donor_utils,
    medical_history as mh_utils,
)
from dcicutils import (
    ff_utils,
    misc_utils,
    schema_utils
)
import pandas as pd
import pprint

logger = logging.getLogger(__name__)

ITEM_TYPES = [
    "Donor",
    "Demographic",
    "DeathCircumstances",
    "MedicalHistory",
    "TissueCollection",
    "FamilyHistory",
    "MedicalTreatment",
    "Diagnosis",
    "Exposure"
]

IGNORED_PROPERTIES = [
    "accession",
    "uuid",
    "tags",
    "submitted_id",
    "date_created",
    "submitted_by",
    "status",
    "schema_version",
    "last_modified",
    "submission_centers",
    "consortia",
    "alternate_accessions",
    "protocols"
]

CHANGED_COLUMNS = {
    "DeathCircumstances.death_pronounced_interval": "DeathCircumstances.death_pronounced_interval_h",
    "DeathCircumstances.ventilator_time": "DeathCircumstances.ventilator_time_h",
    "MedicalHistory.height": "MedicalHistory.height_m",
    "MedicalHistory.weight": "MedicalHistory.weight_kg",
    "TissueCollection.ischemic_time": "TissueCollection.ischemic_time_h",
    "TissueCollection.recovery_interval": "TissueCollection.recovery_interval_min",
    "TissueCollection.refrigeration_prior_to_procurement_time": "TissueCollection.refrigeration_prior_to_procurement_time_h",
    "Exposure.cessation_duration": "Exposure.cessation_duration_y",
    "Exposure.duration": "Exposure.duration_y",
}


def create_bulk_donor_manifest(
    search: str,
    output: str,
    auth_key: Dict[str, str],
    identifiers: Optional[List[str]] = None,
) -> None:
    """Create bulk donor manifest file for given donors."""

    request_handler = RequestHandler(auth_key=auth_key)
    donors = get_donors(search, request_handler, identifiers)
    logger.info(f"Found {len(donors)} donors to process")
    schemas = ff_utils.get_schemas(key=auth_key)
    bulk_donor_manifest = get_bulk_donor_manifest(donors, schemas, request_handler)
    logger.info("Generated bulk donor manifest")
    #log_bulk_donor_manifest(bulk_donor_manifest)
    logger.info(f"Writing out bulk donor manifest to {output}")
    #write_bulk_donor_manfiest(bulk_donor_manifest, output)

def get_donors(
    search: str,
    request_handler: RequestHandler,
    identifiers: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """Get donor items from given search query and idenitfiers."""
    if identifiers:
        return get_donors_from_search(
            search, request_handler.auth_key
        ) + get_donors_from_identifiers(identifiers, request_handler)
    return get_donors_from_search(
            search, request_handler.auth_key
        )

def get_donors_from_search(search_query: str, auth_key: Dict[str, str]) -> List[str]:
    """Get donor items from given search query."""
    if search_query:
        donors = get_items_from_search_query(search_query, auth_key)
        return filter_donors(donors)
    return []


def get_items_from_search_query(
    search_query: str, auth_key: Dict[str, str]
) -> List[str]:
    """Get items from given search query."""
    try:
        result = ff_utils.search_metadata(search_query, key=auth_key)
    except Exception as e:
        logger.error(f"Error searching for items: {e}")
        result = []
    return result


def filter_donors(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Get donors from given items."""
    return [item for item in items if donor_utils.is_donor(item)]


def get_donors_from_identifiers(
    identifiers: List[str], request_handler: RequestHandler
) -> List[str]:
    """Get donors items from given identifiers."""
    items = request_handler.get_items(identifiers)
    return filter_donors(items)


def get_bulk_donor_manifest(
        donors: List[Dict[str, Any]],
        schemas: Dict[str, Any],
        request_handler: RequestHandler
    ) -> pd.DataFrame:
        """Generate dataframe of bulk donor manifest from list of donors."""
        kept_properties = get_kept_properties(schemas)
        donor_manifest = pd.DataFrame(columns=kept_properties)
        external_ids = [item_utils.get_external_id(donor) for donor in donors]
        medical_histories = get_medical_histories(external_ids, request_handler)
        for d in range(1,len(donors)):
            donor = donors[d]
            medical_history = medical_histories[d]
            generate_manifest_row(donor_manifest, d, donor, medical_history, kept_properties)


def get_kept_properties(schemas: Dict[str, Any]):
    """Get properties that are included in the bulk manifest from the schema."""
    all_kept_properties = []
    for item_type in ITEM_TYPES:
        kept_properties = []
        if item_type == "Donor":
            kept_properties+= [f"{item_type}.accession"]
        kept_properties  += [f"{item_type}.{prop}" for prop in schemas[item_type]['properties'].keys() if prop not in IGNORED_PROPERTIES and 'calculatedProperty' not in schemas[item_type]['properties'][prop] and 'linkTo' not in schemas[item_type]['properties'][prop]]
        all_kept_properties+=kept_properties
    modified_properties = [CHANGED_COLUMNS[prop] if prop in CHANGED_COLUMNS.keys() else prop for prop in all_kept_properties ]
    return modified_properties



def get_medical_histories(
        external_ids: List[str],
        request_handler: RequestHandler
) -> List[Dict[str, Any]]:
    """Get medical history list from protected donors."""
    mh_list = []
    search_query = "search/?type=MedicalHistory&frame=embedded"
    for external_id in external_ids:
        search_query += f"&donor={external_id}"
    return order_items_by_donor(mh_list,external_ids)


def order_items_by_donor(items: List[Dict[str, Any]], donor_order: List[str]):
    """Reorder list of item properties to match donor order."""
    new_items = []
    for id in donor_order:
        for item in items:
            if item_utils.get_display_title(
                mh_utils.get_donor(item)
            ) == id:
                new_items.append(item)
    return new_items


def generate_manifest_row(donor_manifest: pd.DataFrame, d: int, donor: str, medical_history: Dict[str, Any], kept_properties: List[str], request_handler: RequestHandler):
    """Generate row for manifest."""
    donor_search_dict = {
        "Donor": f"search/?type=Donor&external_id={donor}&frame=raw",
        "MedicalHistory": f"search/?type=MedicalHistory&donor={donor}&frame=raw",
        "Demographic": f"search/?type=Demographic&donor={donor}&frame=raw",
        "DeathCircumstances": f"search/?type=DeathCircumstances&donor={donor}&frame=raw",
        "TissueCollection": f"search/?type=TissueCollection&donor={donor}&frame=raw",
        "FamilyHistory": f"search/?type=FamilyHistory&donor={donor}&frame=raw",
    }
    mh_search_dict = {
        "Exposure": f"search/?type=Exposure&medical_history={medical_history}&frame=raw",
        "Diagnosis": f"search/?type=Diagnosis&medical_history={medical_history}&frame=raw",
        "MedicalTreatment": f"search/?type=MedicalTreatment&medical_history={medical_history}&frame=raw",
    }
    donor_manifest = add_row_from_search(donor_manifest, d, donor_search_dict, kept_properties, request_handler)
    #donor_manifest = add_row_from_search(donor_manifest, d, mh_search_dict, kept_properties, request_handler)



def add_row_from_search(
        donor_manifest: pd.DataFrame,
        d: int,
        search_dict: Dict[str, Any],
        kept_properties: List[str],
        request_handler: RequestHandler
    ) -> pd.DataFrame:
    """"""
    for item_type, search in search_dict.items():
        hits = ff_utils.search_metadata(search, key=request_handler.auth_key)
        subcolumns = [column.split(".")[1] for column in kept_properties if column.split(".")[0] == item_type]
        if len(hits) > 1:
            for sub in subcolumns:
                col = f"{item_type}.{sub}"
                if col in CHANGED_COLUMNS.keys():
                    col = CHANGED_COLUMNS[col]
                values = []
                for hit in hits:
                    if sub in hit:
                        value = hit[sub]
                        if type(value) is list:
                            value=";".join(value)
                        if value or value == 0:
                            values.append(str(value))
                        else:
                            values.append("NA")
                    else:
                        values.append("NA")
                donor_manifest.at[d, col] = "|".join(values)   
        elif len(hits) == 1:
            for sub in subcolumns:
                col = f"{item_type}.{sub}"
                if col in CHANGED_COLUMNS.keys():
                    col = CHANGED_COLUMNS[col]
                if sub in hits[0]:
                    value = hits[0][sub]
                    if type(value) is list:
                            value=";".join(value)
                    if value or value == 0:
                        donor_manifest.at[d, col] = value
                    else:
                        donor_manifest.at[d, col] = "NA"
                else:
                    donor_manifest.at[d, col] = "NA"
        else:
            for sub in subcolumns:
                col = f"{item_type}.{sub}"
                if col in CHANGED_COLUMNS.keys():
                    col = CHANGED_COLUMNS[col]
                donor_manifest.at[d, col] = "NA"
        return donor_manifest

# def write_bulk_donor_manifest(
#     output: Path,
#     submission_schemas: Dict[str, Any],
#     request_handler: RequestHandler,
#     example: bool = False
# ) -> None:
#     """Write a single workbook containing all submission spreadsheets."""
#     workbook = openpyxl.Workbook()
#     write_workbook_sheets(
#         workbook, ordered_submission_schemas, request_handler, separate_comments=separate_comments, eqm=eqm, example=example
#     )
#     file_path = Path(output, WORKBOOK_FILENAME)
#     save_workbook(workbook, file_path)
#     if example:
#         log.info(f"Example workbook written to: {file_path}")
#     else:
#         log.info(f"Workbook written to: {file_path}")


# @dataclass(frozen=True)
# class Spreadsheet:
#     item: str
#     properties: List[Property]
#     examples: Optional[List[Dict[str,Any]]] = None


# def get_spreadsheet(item: str, submission_schema: Dict[str, Any]) -> Spreadsheet:
#     """Get spreadsheet information for item."""
#     properties = get_properties(item, submission_schema)
#     return Spreadsheet(
#         item=item,
#         properties=properties,
#     )


# def get_properties(item: str, submission_schema: Dict[str, Any]) -> List[Property]:
#     """Get property information from the submission schema"""
#     properties = schema_utils.get_properties(submission_schema)
#     property_list = []
#     for key, value in properties.items():
#         property_list += get_nested_properties(item, key, value)
#     return property_list


# def write_spreadsheet(
#     output: Path, spreadsheet: Spreadsheet, separate_comments: bool = False, example: bool = False
# ) -> None:
#     """Write spreadsheet to file"""
#     file_path = get_output_file_path(output, spreadsheet)
#     workbook = generate_workbook(spreadsheet, separate_comments=separate_comments)
#     save_workbook(workbook, file_path)
#     if example:
#         log.info(f"Example spreadsheet written to: {file_path}")
#     else:
#         log.info(f"Spreadsheet written to: {file_path}")
    

# def get_output_file_path(output: Path, spreadsheet: Spreadsheet) -> Path:
#     """Get the output file path"""
#     return Path(output, f"{to_snake_case(spreadsheet.item)}{ITEM_SPREADSHEET_SUFFIX}")


# def get_property(item: str, property_name: str, property_schema: Dict[str, Any],is_nested: bool = False) -> Property:
#     """Get property information"""
#     return Property(
#         name=property_name,
#         item=item,
#         description=schema_utils.get_description(property_schema),
#         value_type=schema_utils.get_schema_type(property_schema),
#         required=is_required(property_schema),
#         link=is_link(property_schema),
#         enum=get_enum(property_schema),
#         array_subtype=get_array_subtype(property_schema),
#         pattern=schema_utils.get_pattern(property_schema),
#         comment=schema_utils.get_submission_comment(property_schema),
#         examples=get_examples(property_schema),
#         format_=schema_utils.get_format(property_schema),
#         requires=get_corequirements(property_schema),
#         exclusive_requirements=get_exclusive_requirements(property_schema),
#         nested=is_nested,
#         allow_commas=is_allow_commas(property_schema),
#         allow_multiplier_suffix=is_allow_multiplier_suffix(property_schema),
#         search=get_search_url(property_schema)
#     )

# def generate_workbook(
#     spreadsheet: Spreadsheet, separate_comments: bool = False
# ) -> openpyxl.Workbook:
#     """Generate the workbook"""
#     workbook = openpyxl.Workbook()
#     worksheet = workbook.active
#     set_sheet_name(worksheet, spreadsheet)
#     write_properties(
#         worksheet, spreadsheet.properties, separate_comments=separate_comments,examples=spreadsheet.examples
#     )
#     return workbook


# def write_property(
#     worksheet: openpyxl.worksheet.worksheet.Worksheet,
#     index: int,
#     property_: Property,
#     comments: bool = True,
# ) -> None:
#     """Write property to the worksheet"""
#     row = 1  # cells 1-indexed
#     cell = worksheet.cell(row=row, column=index, value=property_.name)
#     set_cell_font(cell, property_)
#     set_cell_width(worksheet, index, property_)
#     if comments:
#         write_comment(worksheet, index, property_)


# def save_workbook(workbook: openpyxl.Workbook, file_path: Path) -> None:
#     """Save the workbook to the file path"""
#     workbook.save(filename=file_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--search",
        "-s",
        help="Search query for donors to create bulk donor manifest",
        default="",
    )
    parser.add_argument(
        "--donors",
        "-d",
        nargs="*",
        help="Donor identifiers to create bulk donor manifest",
        default=[],
    )
    parser.add_argument(
        "--env",
        "-e",
        help="Environment from keys file",
    )
    parser.add_argument(
        "--output",
        "-o",
        help="Output file name",
    )
    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        default=False,
        help="Increase logging verbosity",
    )
    args = parser.parse_args()
    if args.verbose:
        logger.setLevel(logging.DEBUG)
    auth_key = get_auth_key(args.env)
    if not args.search and not args.donors:
        logger.error("Must provide either --search or --identifiers")
    else:
        create_bulk_donor_manifest(
            args.search,
            args.output,
            auth_key,
            args.donors,
        )


if __name__ == "__main__":
    main()